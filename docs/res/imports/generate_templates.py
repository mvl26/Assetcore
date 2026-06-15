"""
AssetCore Import Template Generator
Generates Excel (.xlsx) templates for bulk data import.
Run with: /home/miyano/frappe-bench/env/bin/python generate_templates.py
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = os.path.normpath(
    os.path.join(os.path.dirname(os.path.abspath(__file__)),
                 "..", "..", "assetcore", "public", "import_templates")
)

# ── Color palette ────────────────────────────────────────────────────────────
C_HEADER_REQ  = "C00000"   # dark red   – required field header
C_HEADER_OPT  = "2E4053"   # dark navy  – optional field header
C_HEADER_INFO = "1A5276"   # blue       – info/instruction header
C_HEADER_FG   = "FFFFFF"   # white text
C_ROW_DESC    = "D6EAF8"   # light blue – description row
C_ROW_EXAMPLE = "EAFAF1"   # light green – example row
C_ROW_DATA    = "FFFFFF"   # white – data rows
C_BORDER      = "BDC3C7"

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def header_font(color=C_HEADER_FG, bold=True, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)


def cell_font(bold=False, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_dv(ws, formula, sqref):
    """Add list data validation (dropdown)."""
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Giá trị không hợp lệ",
        error="Vui lòng chọn từ danh sách.",
    )
    dv.sqref = sqref
    ws.add_data_validation(dv)
    return dv


def build_sheet(wb, title, columns, examples, instructions=""):
    """
    columns: list of dict {
        name: str,           # fieldname (EN)
        label: str,          # label (VI)
        desc: str,           # description (VI)
        required: bool,
        example: str,
        dv: str | None,      # dropdown formula e.g. '"A,B,C"'
        width: int,
    }
    examples: list of dict  – actual example data rows (fieldname → value)
    instructions: str       – shown on row 1
    """
    ws = wb.create_sheet(title)

    # ── Row 1: instruction banner ────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1)
    cell.value = (
        f"📋 HƯỚNG DẪN IMPORT: {title}  |  "
        + (instructions or "Điền dữ liệu từ hàng 5 trở xuống. Cột đỏ = bắt buộc. Không xóa hàng 2-4.")
    )
    cell.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    cell.fill = fill(C_HEADER_INFO)
    cell.alignment = left()
    cell.border = BORDER

    # ── Row 2: Field Name (EN) header ────────────────────────────────────
    ws.row_dimensions[2].height = 20
    for col_idx, col in enumerate(columns, start=1):
        c = ws.cell(row=2, column=col_idx)
        c.value = col["name"]
        c.font = header_font()
        c.fill = fill(C_HEADER_REQ if col.get("required") else C_HEADER_OPT)
        c.alignment = center()
        c.border = BORDER

    # ── Row 3: Vietnamese Label ───────────────────────────────────────────
    ws.row_dimensions[3].height = 22
    for col_idx, col in enumerate(columns, start=1):
        c = ws.cell(row=3, column=col_idx)
        req_mark = " (*)" if col.get("required") else ""
        c.value = col["label"] + req_mark
        c.font = Font(name="Calibri", bold=True, size=10,
                      color=("C00000" if col.get("required") else "1A237E"))
        c.fill = fill("FFF9C4")  # light yellow
        c.alignment = center()
        c.border = BORDER

    # ── Row 4: Description ────────────────────────────────────────────────
    ws.row_dimensions[4].height = 42
    for col_idx, col in enumerate(columns, start=1):
        c = ws.cell(row=4, column=col_idx)
        c.value = col.get("desc", "")
        c.font = cell_font(size=9, color="555555")
        c.fill = fill(C_ROW_DESC)
        c.alignment = Alignment(horizontal="left", vertical="top",
                                wrap_text=True)
        c.border = BORDER

    # ── Row 5: Example row ────────────────────────────────────────────────
    ws.row_dimensions[5].height = 20
    example_data = examples[0] if examples else {}
    for col_idx, col in enumerate(columns, start=1):
        c = ws.cell(row=5, column=col_idx)
        c.value = example_data.get(col["name"], col.get("example", ""))
        c.font = cell_font(size=10, color="1B5E20")
        c.fill = fill(C_ROW_EXAMPLE)
        c.alignment = left()
        c.border = BORDER

    # ── Rows 6..105: blank data rows + dropdown validations ───────────────
    for row_idx in range(6, 106):
        ws.row_dimensions[row_idx].height = 18
        for col_idx, col in enumerate(columns, start=1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.fill = fill(C_ROW_DATA if row_idx % 2 == 0 else "F8F9FA")
            c.alignment = left()
            c.border = BORDER
            c.font = cell_font(size=10)

    # Add dropdowns
    for col_idx, col in enumerate(columns, start=1):
        if col.get("dv"):
            col_letter = get_column_letter(col_idx)
            add_dv(ws, col["dv"], f"{col_letter}5:{col_letter}105")

    # ── Column widths ─────────────────────────────────────────────────────
    for col_idx, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col.get("width", 20)

    # Freeze panes below row 4
    ws.freeze_panes = "A5"
    return ws


# ════════════════════════════════════════════════════════════════════════════
# 1a. DANH MỤC TÀI SẢN (AC Asset Category)
# ════════════════════════════════════════════════════════════════════════════

def make_asset_category():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Danh mục tài sản (Category)", [
        {"name": "category_name", "label": "Tên danh mục", "required": True, "width": 30,
         "desc": "Tên duy nhất của danh mục thiết bị y tế.\nVD: Máy chẩn đoán hình ảnh",
         "example": "Máy chẩn đoán hình ảnh"},
        {"name": "description", "label": "Mô tả", "required": False, "width": 40,
         "desc": "Mô tả ngắn về danh mục.",
         "example": "Các thiết bị siêu âm, X-quang, MRI, CT"},
        {"name": "gmdn_code", "label": "Mã GMDN", "required": False, "width": 15,
         "desc": "Mã Global Medical Device Nomenclature (nếu có).",
         "example": "57947"},
        {"name": "gmdn_term", "label": "Tên GMDN", "required": False, "width": 35,
         "desc": "Thuật ngữ GMDN tương ứng.",
         "example": "Diagnostic ultrasound scanner"},
        {"name": "default_pm_required", "label": "Cần bảo trì định kỳ?", "required": False, "width": 22,
         "desc": "1 = có; 0 = không.\nMặc định khi tạo thiết bị mới.", "example": "1",
         "dv": '"1,0"'},
        {"name": "default_pm_interval_days", "label": "Chu kỳ bảo trì (ngày)", "required": False, "width": 24,
         "desc": "Số ngày giữa hai lần bảo trì. Bắt buộc nếu PM = 1.",
         "example": "180"},
        {"name": "default_calibration_required", "label": "Cần hiệu chuẩn?", "required": False, "width": 20,
         "desc": "1 = có; 0 = không.", "example": "1",
         "dv": '"1,0"'},
        {"name": "default_calibration_interval_days", "label": "Chu kỳ hiệu chuẩn (ngày)", "required": False, "width": 26,
         "desc": "Số ngày giữa hai lần hiệu chuẩn.",
         "example": "365"},
        {"name": "default_depreciation_method", "label": "Phương pháp khấu hao", "required": False, "width": 26,
         "desc": "Straight Line / Double Declining / Units of Production",
         "example": "Straight Line",
         "dv": '"Straight Line,Double Declining,Units of Production"'},
        {"name": "total_depreciation_months", "label": "Thời gian khấu hao (tháng)", "required": False, "width": 28,
         "desc": "Tổng số tháng khấu hao. VD: 60 = 5 năm.", "example": "60"},
        {"name": "has_radiation", "label": "Có bức xạ?", "required": False, "width": 16,
         "desc": "1 = thiết bị bức xạ (X-quang, CT, PET...); 0 = không.",
         "example": "0", "dv": '"1,0"'},
        {"name": "is_active", "label": "Đang hoạt động", "required": False, "width": 18,
         "desc": "1 = đang dùng; 0 = vô hiệu. Mặc định: 1.",
         "example": "1", "dv": '"1,0"'},
    ], [{}])

    wb.save(os.path.join(OUT_DIR, "01a_danh_muc_tai_san.xlsx"))
    print("✓ 01a_danh_muc_tai_san.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 1b. KHOA PHÒNG (AC Department)
# ════════════════════════════════════════════════════════════════════════════

def make_department():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Khoa phòng (Department)", [
        {"name": "department_name", "label": "Tên khoa/phòng", "required": True, "width": 35,
         "desc": "Tên đầy đủ của khoa hoặc phòng ban.",
         "example": "Khoa Hồi sức tích cực"},
        {"name": "department_code", "label": "Mã khoa/phòng", "required": False, "width": 18,
         "desc": "Mã ngắn duy nhất. VD: ICU, XRAY, LAB.",
         "example": "ICU"},
        {"name": "parent_department", "label": "Khoa cha", "required": False, "width": 30,
         "desc": "Tên khoa cha (nếu là khoa con). Phải tồn tại trong hệ thống.",
         "example": "Bệnh viện ABC"},
        {"name": "dept_head", "label": "Trưởng khoa (email)", "required": False, "width": 30,
         "desc": "Email của trưởng khoa (User phải tồn tại trong hệ thống).",
         "example": "nguyen.van.a@hospital.vn"},
        {"name": "phone", "label": "Số điện thoại", "required": False, "width": 18,
         "desc": "Số DT nội bộ hoặc di động.", "example": "0912345678"},
        {"name": "email", "label": "Email khoa", "required": False, "width": 30,
         "desc": "Email liên hệ của khoa.", "example": "icu@hospital.vn"},
        {"name": "is_group", "label": "Là nhóm cha?", "required": False, "width": 16,
         "desc": "1 = node cha (không phải lá); 0 = khoa thực.",
         "example": "0", "dv": '"1,0"'},
        {"name": "is_active", "label": "Đang hoạt động", "required": False, "width": 18,
         "desc": "1 = đang dùng; 0 = vô hiệu.", "example": "1", "dv": '"1,0"'},
    ], [{}])

    wb.save(os.path.join(OUT_DIR, "01b_khoa_phong.xlsx"))
    print("✓ 01b_khoa_phong.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 1c. VỊ TRÍ (AC Location)
# ════════════════════════════════════════════════════════════════════════════

def make_location():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Vị trí (Location)", [
        {"name": "location_name", "label": "Tên vị trí", "required": True, "width": 35,
         "desc": "Tên đầy đủ vị trí/phòng đặt thiết bị.",
         "example": "Phòng hồi sức P301"},
        {"name": "location_code", "label": "Mã vị trí", "required": False, "width": 18,
         "desc": "Mã ngắn duy nhất. VD: P301, XRAY-1.",
         "example": "P301"},
        {"name": "parent_location", "label": "Vị trí cha", "required": False, "width": 30,
         "desc": "Tên vị trí cha (phải tồn tại). VD: Tòa nhà A.",
         "example": "Tòa nhà A"},
        {"name": "clinical_area_type", "label": "Loại khu vực lâm sàng", "required": False, "width": 26,
         "desc": "ICU / OR / Lab / Imaging / General Ward / Storage / Office",
         "example": "ICU",
         "dv": '"ICU,OR,Lab,Imaging,General Ward,Storage,Office"'},
        {"name": "infection_control_level", "label": "Mức kiểm soát nhiễm khuẩn", "required": False, "width": 30,
         "desc": "Standard / Enhanced / Isolation",
         "example": "Standard",
         "dv": '"Standard,Enhanced,Isolation"'},
        {"name": "power_backup_available", "label": "Có UPS/máy phát?", "required": False, "width": 20,
         "desc": "1 = có nguồn dự phòng; 0 = không.",
         "example": "1", "dv": '"1,0"'},
        {"name": "dept_head", "label": "Người phụ trách (email)", "required": False, "width": 30,
         "desc": "Email User phụ trách vị trí.", "example": "nguyen.van.a@hospital.vn"},
        {"name": "contact_phone", "label": "Số liên hệ", "required": False, "width": 22,
         "desc": "Số liên hệ. Khi để trống, hệ thống sẽ tự lấy theo số di động "
                 "của người phụ trách (mobile_no).",
         "example": "0901234567"},
        {"name": "is_group", "label": "Là nhóm cha?", "required": False, "width": 16,
         "desc": "1 = node cha; 0 = vị trí thực.", "example": "0", "dv": '"1,0"'},
    ], [{}])

    wb.save(os.path.join(OUT_DIR, "01c_vi_tri.xlsx"))
    print("✓ 01c_vi_tri.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 2. IMM-00 — NCC / Model / Hợp đồng / SLA
# ════════════════════════════════════════════════════════════════════════════

def make_imm00():
    wb = Workbook()
    wb.remove(wb.active)

    # ── 2a. AC Supplier ───────────────────────────────────────────────────
    build_sheet(wb, "Nhà cung cấp (Supplier)", [
        {"name": "supplier_name", "label": "Tên NCC", "required": True, "width": 35,
         "desc": "Tên đầy đủ của nhà cung cấp / nhà thầu.",
         "example": "Công ty TNHH Thiết bị Y tế ABC"},
        {"name": "supplier_code", "label": "Mã NCC", "required": False, "width": 15,
         "desc": "Mã nội bộ duy nhất. Tự sinh nếu để trống.", "example": "SUP-001"},
        {"name": "supplier_group", "label": "Loại NCC", "required": True, "width": 22,
         "desc": "Manufacturer / Distributor / Calibration Lab / Service Provider",
         "example": "Distributor",
         "dv": '"Manufacturer,Distributor,Calibration Lab,Service Provider"'},
        {"name": "vendor_type", "label": "Loại vendor", "required": True, "width": 22,
         "desc": "Manufacturer / Distributor / Calibration Lab / Service",
         "example": "Distributor",
         "dv": '"Manufacturer,Distributor,Calibration Lab,Service"'},
        {"name": "country", "label": "Quốc gia", "required": False, "width": 16,
         "desc": "Quốc gia đăng ký kinh doanh. VD: Việt Nam, Singapore.",
         "example": "Việt Nam"},
        {"name": "tax_id", "label": "Mã số thuế", "required": False, "width": 18,
         "desc": "MST doanh nghiệp (10-13 chữ số).", "example": "0123456789"},
        {"name": "address", "label": "Địa chỉ", "required": False, "width": 40,
         "desc": "Địa chỉ trụ sở chính.", "example": "123 Nguyễn Huệ, Q.1, TP.HCM"},
        {"name": "phone", "label": "Điện thoại", "required": False, "width": 18,
         "desc": "Số điện thoại công ty.", "example": "028 3822 1234"},
        {"name": "email_id", "label": "Email", "required": False, "width": 30,
         "desc": "Email liên hệ chính.", "example": "info@abc-medical.vn"},
        {"name": "support_hotline", "label": "Hotline hỗ trợ", "required": False, "width": 20,
         "desc": "Số hotline hỗ trợ kỹ thuật 24/7.", "example": "1800 1234"},
        {"name": "technical_email", "label": "Email kỹ thuật", "required": False, "width": 30,
         "desc": "Email bộ phận kỹ thuật.", "example": "support@abc-medical.vn"},
        {"name": "local_representative", "label": "Đại diện tại VN", "required": False, "width": 28,
         "desc": "Tên người đại diện tại Việt Nam.", "example": "Nguyễn Văn B"},
        {"name": "website", "label": "Website", "required": False, "width": 28,
         "desc": "Website công ty.", "example": "https://abc-medical.vn"},
        {"name": "iso_17025_cert", "label": "Chứng chỉ ISO 17025", "required": False, "width": 24,
         "desc": "Số chứng chỉ ISO/IEC 17025 (nếu là phòng hiệu chuẩn).", "example": "VILAS-123"},
        {"name": "iso_17025_expiry", "label": "Hết hạn ISO 17025", "required": False, "width": 22,
         "desc": "Ngày hết hạn ISO 17025. Định dạng: YYYY-MM-DD.", "example": "2026-12-31"},
        {"name": "iso_13485_cert", "label": "Chứng chỉ ISO 13485", "required": False, "width": 24,
         "desc": "Số chứng chỉ ISO 13485 (thiết bị y tế).", "example": "ISO13485-456"},
        {"name": "iso_13485_expiry", "label": "Hết hạn ISO 13485", "required": False, "width": 22,
         "desc": "Ngày hết hạn ISO 13485. Định dạng: YYYY-MM-DD.", "example": "2027-06-30"},
        {"name": "is_active", "label": "Đang hoạt động", "required": False, "width": 18,
         "desc": "1 = đang hợp tác; 0 = vô hiệu.", "example": "1", "dv": '"1,0"'},
    ], [{}])

    # ── 2b. IMM Device Model ──────────────────────────────────────────────
    build_sheet(wb, "Mô hình thiết bị (Device Model)", [
        {"name": "model_name", "label": "Tên model", "required": True, "width": 35,
         "desc": "Tên model chính thức của nhà sản xuất.", "example": "LOGIQ E10"},
        {"name": "manufacturer", "label": "Nhà sản xuất", "required": True, "width": 28,
         "desc": "Tên nhà sản xuất (hãng gốc).", "example": "GE Healthcare"},
        {"name": "asset_category", "label": "Danh mục tài sản", "required": True, "width": 28,
         "desc": "Phải khớp với tên trong sheet 'Danh mục tài sản'.",
         "example": "Máy chẩn đoán hình ảnh"},
        {"name": "medical_device_class", "label": "Phân loại thiết bị", "required": True, "width": 22,
         "desc": "Class I / Class II / Class III (theo QCVN/WHO).",
         "example": "Class II",
         "dv": '"Class I,Class II,Class III"'},
        {"name": "model_version", "label": "Phiên bản", "required": False, "width": 16,
         "desc": "Số phiên bản / generation. VD: Rev.3, Gen2.", "example": "S8"},
        {"name": "country_of_origin", "label": "Xuất xứ", "required": False, "width": 16,
         "desc": "Quốc gia sản xuất.", "example": "Mỹ"},
        {"name": "expected_lifespan_years", "label": "Tuổi thọ kỳ vọng (năm)", "required": False, "width": 26,
         "desc": "Tuổi thọ trung bình do nhà sản xuất khuyến nghị.", "example": "10"},
        {"name": "gmdn_code", "label": "Mã GMDN", "required": False, "width": 14,
         "desc": "Mã định danh GMDN quốc tế.", "example": "57947"},
        {"name": "gmdn_term", "label": "Tên GMDN", "required": False, "width": 35,
         "desc": "Tên GMDN tiếng Anh.", "example": "Diagnostic ultrasound scanner"},
        {"name": "registration_required", "label": "Cần số đăng ký BYT?", "required": False, "width": 24,
         "desc": "1 = bắt buộc có số ĐKLH Bộ Y tế; 0 = không.", "example": "1", "dv": '"1,0"'},
        {"name": "is_radiation_device", "label": "Thiết bị bức xạ?", "required": False, "width": 20,
         "desc": "1 = có phát xạ ion hóa (X-quang, CT...); 0 = không.", "example": "0", "dv": '"1,0"'},
        {"name": "is_pm_required", "label": "Cần bảo trì ĐK?", "required": False, "width": 20,
         "desc": "1 = có lịch PM; 0 = không.", "example": "1", "dv": '"1,0"'},
        {"name": "pm_interval_days", "label": "Chu kỳ PM (ngày)", "required": False, "width": 20,
         "desc": "Số ngày giữa 2 lần bảo trì. Bắt buộc nếu PM = 1.", "example": "180"},
        {"name": "pm_alert_days", "label": "Cảnh báo PM trước (ngày)", "required": False, "width": 26,
         "desc": "Số ngày cảnh báo trước khi đến hạn PM. Mặc định: 30.", "example": "30"},
        {"name": "is_calibration_required", "label": "Cần hiệu chuẩn?", "required": False, "width": 20,
         "desc": "1 = có lịch hiệu chuẩn; 0 = không.", "example": "1", "dv": '"1,0"'},
        {"name": "calibration_interval_days", "label": "Chu kỳ hiệu chuẩn (ngày)", "required": False, "width": 28,
         "desc": "Số ngày giữa 2 lần hiệu chuẩn. Bắt buộc nếu HC = 1.", "example": "365"},
        {"name": "default_calibration_type", "label": "Loại hiệu chuẩn", "required": False, "width": 22,
         "desc": "Internal / External / Both",
         "example": "External",
         "dv": '"Internal,External,Both"'},
        {"name": "power_supply", "label": "Nguồn điện", "required": False, "width": 20,
         "desc": "VD: 220V/50Hz, 110-240V/50-60Hz.", "example": "220V/50Hz"},
        {"name": "dimensions", "label": "Kích thước (DxRxC, mm)", "required": False, "width": 26,
         "desc": "Dài x Rộng x Cao, đơn vị mm.", "example": "600x800x1200"},
        {"name": "weight_kg", "label": "Trọng lượng (kg)", "required": False, "width": 20,
         "desc": "Trọng lượng máy (kg).", "example": "85.5"},
        {"name": "specifications", "label": "Thông số kỹ thuật", "required": False, "width": 40,
         "desc": "Mô tả thông số kỹ thuật chính.", "example": "Frequency: 1-18 MHz; 4 probe ports"},
    ], [{}])

    # ── 2c. Service Contract ──────────────────────────────────────────────
    build_sheet(wb, "Hợp đồng (Service Contract)", [
        {"name": "contract_code", "label": "Số hợp đồng", "required": True, "width": 22,
         "desc": "Số hợp đồng duy nhất. VD: HD-2024-001.", "example": "HD-2024-001"},
        {"name": "contract_title", "label": "Tên hợp đồng", "required": True, "width": 40,
         "desc": "Tên đầy đủ của hợp đồng.", "example": "Hợp đồng bảo trì thiết bị siêu âm 2024"},
        {"name": "supplier", "label": "Nhà cung cấp", "required": True, "width": 35,
         "desc": "Tên NCC phải khớp với sheet 'Nhà cung cấp'.", "example": "Công ty TNHH Thiết bị Y tế ABC"},
        {"name": "contract_type", "label": "Loại hợp đồng", "required": True, "width": 25,
         "desc": "Preventive Maintenance / Calibration / Repair / Full Service / Warranty Extension",
         "example": "Preventive Maintenance",
         "dv": '"Preventive Maintenance,Calibration,Repair,Full Service,Warranty Extension"'},
        {"name": "contract_start", "label": "Ngày bắt đầu", "required": True, "width": 18,
         "desc": "Định dạng: YYYY-MM-DD.", "example": "2024-01-01"},
        {"name": "contract_end", "label": "Ngày kết thúc", "required": True, "width": 18,
         "desc": "Định dạng: YYYY-MM-DD.", "example": "2024-12-31"},
        {"name": "sign_date", "label": "Ngày ký", "required": False, "width": 16,
         "desc": "Ngày ký hợp đồng. Định dạng: YYYY-MM-DD.", "example": "2023-12-15"},
        {"name": "contract_value", "label": "Giá trị hợp đồng (VNĐ)", "required": False, "width": 26,
         "desc": "Tổng giá trị hợp đồng (VNĐ, không dấu chấm/phẩy).", "example": "150000000"},
        {"name": "sla_response_hours", "label": "SLA phản hồi (giờ)", "required": False, "width": 22,
         "desc": "Thời gian phản hồi cam kết (giờ). VD: 4 = 4 giờ.", "example": "4"},
        {"name": "auto_renew", "label": "Tự gia hạn?", "required": False, "width": 16,
         "desc": "1 = tự động gia hạn; 0 = không.", "example": "0", "dv": '"1,0"'},
        {"name": "coverage_description", "label": "Phạm vi bảo hiểm/bảo trì", "required": False, "width": 40,
         "desc": "Mô tả nội dung dịch vụ được bao gồm trong hợp đồng.",
         "example": "PM 2 lần/năm, sửa chữa không giới hạn, phụ tùng theo giá gốc"},
        {"name": "notes", "label": "Ghi chú", "required": False, "width": 40,
         "desc": "Ghi chú thêm về hợp đồng.", "example": ""},
    ], [{}])

    # ── 2d. IMM SLA Policy ────────────────────────────────────────────────
    build_sheet(wb, "Chính sách SLA", [
        {"name": "policy_name", "label": "Tên chính sách SLA", "required": True, "width": 35,
         "desc": "Tên duy nhất của chính sách SLA. VD: SLA-Critical-ICU.",
         "example": "SLA-P1-Critical"},
        {"name": "priority", "label": "Mức ưu tiên", "required": True, "width": 18,
         "desc": "P1 Critical / P1 High / P2 / P3 / P4",
         "example": "P1 Critical",
         "dv": '"P1 Critical,P1 High,P2,P3,P4"'},
        {"name": "response_time_minutes", "label": "Thời gian phản hồi (phút)", "required": True, "width": 28,
         "desc": "Thời gian tối đa để ghi nhận / phản hồi sự cố (phút). VD: 30.",
         "example": "30"},
        {"name": "resolution_time_hours", "label": "Thời gian giải quyết (giờ)", "required": True, "width": 28,
         "desc": "Thời gian tối đa để hoàn tất sửa chữa (giờ). VD: 4.",
         "example": "4"},
        {"name": "risk_class", "label": "Phân loại rủi ro", "required": False, "width": 22,
         "desc": "Low / Medium / High / Critical — áp dụng cho rủi ro nào.",
         "example": "Critical",
         "dv": '"Low,Medium,High,Critical"'},
        {"name": "escalation_l1_user", "label": "Escalation L1 (email)", "required": False, "width": 28,
         "desc": "Email User sẽ nhận thông báo leo thang L1.", "example": "truong.kythuat@hospital.vn"},
        {"name": "escalation_l1_hours", "label": "Leo thang L1 sau (giờ)", "required": False, "width": 24,
         "desc": "Số giờ chưa giải quyết thì leo thang L1.", "example": "2"},
        {"name": "escalation_l2_user", "label": "Escalation L2 (email)", "required": False, "width": 28,
         "desc": "Email User sẽ nhận thông báo leo thang L2.", "example": "giam.doc.ky.thuat@hospital.vn"},
        {"name": "escalation_l2_hours", "label": "Leo thang L2 sau (giờ)", "required": False, "width": 24,
         "desc": "Số giờ chưa giải quyết thì leo thang L2.", "example": "8"},
        {"name": "effective_date", "label": "Ngày hiệu lực", "required": False, "width": 18,
         "desc": "Định dạng: YYYY-MM-DD.", "example": "2024-01-01"},
        {"name": "expiry_date", "label": "Ngày hết hiệu lực", "required": False, "width": 20,
         "desc": "Định dạng: YYYY-MM-DD. Để trống = không hết hạn.", "example": ""},
        {"name": "is_default", "label": "SLA mặc định?", "required": False, "width": 18,
         "desc": "1 = áp dụng mặc định khi không có SLA chuyên biệt.", "example": "0", "dv": '"1,0"'},
        {"name": "is_active", "label": "Đang hoạt động", "required": False, "width": 18,
         "desc": "1 = đang dùng; 0 = vô hiệu.", "example": "1", "dv": '"1,0"'},
    ], [{}])

    wb.save(os.path.join(OUT_DIR, "02_imm00_ncc_model_hopdong_sla.xlsx"))
    print("✓ 02_imm00_ncc_model_hopdong_sla.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 3. DANH SÁCH TÀI SẢN (AC Asset)
# ════════════════════════════════════════════════════════════════════════════

def make_assets():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Danh sách tài sản", [
        # ── Thông tin cơ bản ──────────────────────────────────────────────
        {"name": "asset_name", "label": "Tên tài sản", "required": True, "width": 35,
         "desc": "Tên đầy đủ của thiết bị. VD: Máy siêu âm tim mạch phòng Tim.",
         "example": "Máy siêu âm LOGIQ E10 - Khoa Tim mạch"},
        {"name": "asset_code", "label": "Mã tài sản", "required": False, "width": 22,
         "desc": "Mã định danh nội bộ duy nhất. Chỉ chữ, số và . _ - / "
                 "(không khoảng trắng, không dấu). Để trống = hệ thống tự sinh.",
         "example": "TS-2024-001"},
        {"name": "asset_category", "label": "Danh mục tài sản", "required": True, "width": 28,
         "desc": "Phải khớp tên trong file 01_du_lieu_tham_chieu.xlsx / sheet 'Danh mục tài sản'.",
         "example": "Máy chẩn đoán hình ảnh"},
        {"name": "device_model", "label": "Model thiết bị", "required": False, "width": 25,
         "desc": "Tên model trong file 02 / sheet 'Mô hình thiết bị'. Tự điền nếu có.",
         "example": "LOGIQ E10"},
        {"name": "manufacturer_sn", "label": "Số serial NSX", "required": False, "width": 22,
         "desc": "Số serial do nhà sản xuất cấp — KHÁC Mã tài sản. "
                 "Không bắt buộc; nếu nhập phải duy nhất.",
         "example": "SN-GE-2023-88712"},
        {"name": "udi_code", "label": "Mã UDI", "required": False, "width": 28,
         "desc": "Unique Device Identifier theo FDA/MDR (nếu có).", "example": "00850000000000"},
        {"name": "byt_reg_no", "label": "Số ĐKLH Bộ Y tế", "required": False, "width": 26,
         "desc": "Số đăng ký lưu hành do Bộ Y tế cấp.", "example": "TTBYT-2023-12345"},
        {"name": "byt_reg_expiry", "label": "Hết hạn ĐKLH", "required": False, "width": 18,
         "desc": "Ngày hết hạn số ĐKLH. Định dạng: YYYY-MM-DD.", "example": "2028-06-30"},
        # ── Vị trí & phụ trách ───────────────────────────────────────────
        {"name": "location", "label": "Vị trí đặt máy", "required": False, "width": 28,
         "desc": "Tên vị trí (khớp file 01 / sheet 'Vị trí'). VD: Phòng hồi sức P301.",
         "example": "Phòng siêu âm tim - Tầng 3"},
        {"name": "department", "label": "Khoa phòng quản lý", "required": False, "width": 28,
         "desc": "Tên khoa phòng (khớp file 01 / sheet 'Khoa phòng').",
         "example": "Khoa Tim mạch"},
        {"name": "custodian", "label": "Người phụ trách (email)", "required": False, "width": 30,
         "desc": "Email của người giữ/sử dụng thiết bị.", "example": "nguyen.van.a@hospital.vn"},
        {"name": "responsible_technician", "label": "KTV phụ trách (email)", "required": False, "width": 30,
         "desc": "Email kỹ thuật viên chịu trách nhiệm bảo trì.", "example": "le.van.b@hospital.vn"},
        {"name": "supplier", "label": "Nhà cung cấp", "required": False, "width": 30,
         "desc": "Tên NCC (khớp file 02 / sheet 'Nhà cung cấp').",
         "example": "Công ty TNHH Thiết bị Y tế ABC"},
        # ── Mua sắm & bảo hành ───────────────────────────────────────────
        {"name": "purchase_date", "label": "Ngày mua", "required": False, "width": 16,
         "desc": "Định dạng: YYYY-MM-DD.", "example": "2022-03-15"},
        {"name": "gross_purchase_amount", "label": "Giá mua (VNĐ)", "required": False, "width": 20,
         "desc": "Nguyên giá tài sản khi mua (VNĐ, không dấu phẩy).", "example": "450000000"},
        {"name": "warranty_expiry_date", "label": "Hết hạn bảo hành", "required": False, "width": 22,
         "desc": "Ngày hết hạn bảo hành NSX. Định dạng: YYYY-MM-DD.", "example": "2025-03-15"},
        {"name": "commissioning_date", "label": "Ngày đưa vào sử dụng", "required": False, "width": 26,
         "desc": "Ngày nghiệm thu / đưa vào vận hành. Định dạng: YYYY-MM-DD.", "example": "2022-04-01"},
        # ── Khấu hao ─────────────────────────────────────────────────────
        {"name": "depreciation_method", "label": "Phương pháp khấu hao", "required": False, "width": 26,
         "desc": "Straight Line / Double Declining / Units of Production",
         "example": "Straight Line",
         "dv": '"Straight Line,Double Declining,Units of Production"'},
        {"name": "useful_life_years", "label": "Tuổi thọ hữu ích (năm)", "required": False, "width": 26,
         "desc": "Số năm khấu hao theo quy định tài chính.", "example": "10"},
        {"name": "residual_value", "label": "Giá trị thu hồi (VNĐ)", "required": False, "width": 24,
         "desc": "Giá trị còn lại ước tính cuối vòng đời (VNĐ).", "example": "45000000"},
        {"name": "in_service_date", "label": "Ngày bắt đầu khấu hao", "required": False, "width": 26,
         "desc": "Ngày bắt đầu tính khấu hao. Định dạng: YYYY-MM-DD.", "example": "2022-04-01"},
        # ── Bảo hiểm ─────────────────────────────────────────────────────
        {"name": "insurance_policy_no", "label": "Số hợp đồng bảo hiểm", "required": False, "width": 26,
         "desc": "Số hợp đồng bảo hiểm tài sản (nếu có).", "example": "BH-2024-55678"},
        {"name": "insurer_name", "label": "Công ty bảo hiểm", "required": False, "width": 26,
         "desc": "Tên công ty bảo hiểm.", "example": "Bảo Việt"},
        {"name": "insured_value", "label": "Số tiền bảo hiểm (VNĐ)", "required": False, "width": 26,
         "desc": "Giá trị bảo hiểm (VNĐ).", "example": "400000000"},
        {"name": "insurance_start_date", "label": "Ngày bắt đầu BH", "required": False, "width": 22,
         "desc": "Định dạng: YYYY-MM-DD.", "example": "2024-01-01"},
        {"name": "insurance_end_date", "label": "Ngày hết hạn BH", "required": False, "width": 22,
         "desc": "Định dạng: YYYY-MM-DD.", "example": "2024-12-31"},
        # ── Trạng thái ───────────────────────────────────────────────────
        {"name": "lifecycle_status", "label": "Trạng thái vòng đời", "required": True, "width": 26,
         "desc": "Draft / Commissioned / Active / Under Maintenance / Under Repair / Calibrating / Out of Service / Decommissioned",
         "example": "Active",
         "dv": '"Draft,Commissioned,Active,Under Maintenance,Under Repair,Calibrating,Out of Service,Decommissioned"'},
        {"name": "notes", "label": "Ghi chú", "required": False, "width": 40,
         "desc": "Ghi chú thêm về tài sản.", "example": ""},
    ], [{}],
    instructions="Mỗi hàng = 1 tài sản. Cột đỏ bắt buộc. "
                 "Các cột Link phải khớp giá trị đã có trong hệ thống hoặc file import tham chiếu.")

    wb.save(os.path.join(OUT_DIR, "03_danh_sach_tai_san.xlsx"))
    print("✓ 03_danh_sach_tai_san.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 4. PHỤ TÙNG (AC Spare Part)
# ════════════════════════════════════════════════════════════════════════════

def make_spare_parts():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Danh sách phụ tùng", [
        {"name": "part_name", "label": "Tên phụ tùng", "required": True, "width": 35,
         "desc": "Tên đầy đủ của phụ tùng / vật tư.", "example": "Cáp ECG 10 nhánh (GE)"},
        {"name": "part_code", "label": "Mã phụ tùng", "required": False, "width": 20,
         "desc": "Mã nội bộ duy nhất. Để trống = tự sinh.", "example": "PT-ECG-001"},
        {"name": "part_category", "label": "Loại phụ tùng", "required": False, "width": 20,
         "desc": "Electrical / Mechanical / Consumable / Filter / Battery / Sensor / Other",
         "example": "Electrical",
         "dv": '"Electrical,Mechanical,Consumable,Filter,Battery,Sensor,Other"'},
        {"name": "manufacturer", "label": "Nhà sản xuất", "required": False, "width": 25,
         "desc": "Tên hãng sản xuất phụ tùng.", "example": "GE Healthcare"},
        {"name": "manufacturer_part_no", "label": "Mã NSX (Part No.)", "required": False, "width": 22,
         "desc": "Mã phụ tùng theo nhà sản xuất.", "example": "2017716-001"},
        {"name": "preferred_supplier", "label": "NCC ưu tiên", "required": False, "width": 28,
         "desc": "Tên NCC ưu tiên (khớp file 02 / sheet 'Nhà cung cấp').",
         "example": "Công ty TNHH Thiết bị Y tế ABC"},
        {"name": "stock_uom", "label": "Đơn vị tính", "required": True, "width": 16,
         "desc": "Đơn vị tính lưu kho. VD: Cái, Bộ, Hộp, Cuộn.",
         "example": "Cái"},
        {"name": "purchase_uom", "label": "Đơn vị mua", "required": False, "width": 16,
         "desc": "Đơn vị tính khi mua (nếu khác lưu kho).", "example": "Cái"},
        {"name": "unit_cost", "label": "Đơn giá (VNĐ)", "required": False, "width": 18,
         "desc": "Giá mua trung bình (VNĐ, không dấu phẩy).", "example": "2500000"},
        {"name": "min_stock_level", "label": "Tồn kho tối thiểu", "required": False, "width": 22,
         "desc": "Mức tồn kho tối thiểu cần cảnh báo.", "example": "2"},
        {"name": "max_stock_level", "label": "Tồn kho tối đa", "required": False, "width": 20,
         "desc": "Mức tồn kho tối đa (giới hạn nhập).", "example": "10"},
        {"name": "shelf_life_months", "label": "Hạn sử dụng (tháng)", "required": False, "width": 22,
         "desc": "Hạn sử dụng tính bằng tháng (0 = không giới hạn).", "example": "24"},
        {"name": "is_critical", "label": "Phụ tùng thiết yếu?", "required": False, "width": 22,
         "desc": "1 = quan trọng, phải luôn có trong kho; 0 = bình thường.",
         "example": "1", "dv": '"1,0"'},
        {"name": "specifications", "label": "Thông số kỹ thuật", "required": False, "width": 40,
         "desc": "Mô tả thông số, quy cách kỹ thuật của phụ tùng.",
         "example": "10-lead, 4.0m, IEC connector, latex-free"},
        {"name": "is_active", "label": "Đang sử dụng", "required": False, "width": 18,
         "desc": "1 = đang trong danh mục; 0 = ngừng dùng.", "example": "1", "dv": '"1,0"'},
    ], [{}])

    wb.save(os.path.join(OUT_DIR, "04_danh_sach_phu_tung.xlsx"))
    print("✓ 04_danh_sach_phu_tung.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 5. KHO HÀNG (AC Warehouse)
# ════════════════════════════════════════════════════════════════════════════

def make_warehouse():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Kho hàng (Warehouse)", [
        {"name": "warehouse_code", "label": "Mã kho", "required": True, "width": 18,
         "desc": "Mã kho duy nhất. VD: KHO-KT, KHO-VAT-TU.", "example": "KHO-KT-01"},
        {"name": "warehouse_name", "label": "Tên kho", "required": True, "width": 35,
         "desc": "Tên đầy đủ của kho hàng.", "example": "Kho vật tư kỹ thuật trung tâm"},
        {"name": "location", "label": "Vị trí kho", "required": False, "width": 28,
         "desc": "Tên vị trí (khớp file 01 / sheet 'Vị trí').", "example": "Tầng hầm - Tòa nhà A"},
        {"name": "department", "label": "Khoa phòng quản lý", "required": False, "width": 28,
         "desc": "Tên khoa phòng chịu trách nhiệm kho (khớp file 01).",
         "example": "Phòng Vật tư thiết bị"},
        {"name": "manager", "label": "Thủ kho (email)", "required": False, "width": 30,
         "desc": "Email của thủ kho / người phụ trách kho.", "example": "tran.van.c@hospital.vn"},
        {"name": "is_active", "label": "Đang hoạt động", "required": False, "width": 18,
         "desc": "1 = đang dùng; 0 = vô hiệu.", "example": "1", "dv": '"1,0"'},
        {"name": "notes", "label": "Ghi chú", "required": False, "width": 40,
         "desc": "Ghi chú về kho (điều kiện bảo quản, loại vật tư...).",
         "example": "Kho lạnh 2-8°C cho vật tư sinh phẩm"},
    ], [{}])

    wb.save(os.path.join(OUT_DIR, "05_kho_hang.xlsx"))
    print("✓ 05_kho_hang.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 6. USERS (Frappe User)
# ════════════════════════════════════════════════════════════════════════════

def make_users():
    wb = Workbook()
    wb.remove(wb.active)

    build_sheet(wb, "Danh sách người dùng", [
        {"name": "email", "label": "Email (Mã người dùng)", "required": True, "width": 32,
         "desc": "Email là username đăng nhập hệ thống. Phải là email hợp lệ, duy nhất.",
         "example": "nguyen.van.a@hospital.vn"},
        {"name": "first_name", "label": "Tên", "required": True, "width": 18,
         "desc": "Tên (không bao gồm họ).", "example": "Văn A"},
        {"name": "last_name", "label": "Họ", "required": False, "width": 18,
         "desc": "Họ của người dùng.", "example": "Nguyễn"},
        {"name": "mobile_no", "label": "Điện thoại", "required": False, "width": 18,
         "desc": "Số di động. VD: 0912345678.", "example": "0912345678"},
        {"name": "ac_department", "label": "Khoa/Phòng", "required": False, "width": 28,
         "desc": "Mã khoa/phòng (AC Department name). VD: Khoa-HSTC.",
         "example": "Khoa-HSTC"},
        {"name": "imm_approval_status", "label": "Trạng thái duyệt", "required": False, "width": 22,
         "desc": "Trạng thái duyệt IMM. Chọn: Pending / Approved / Rejected.",
         "example": "Approved", "dv": '"Pending,Approved,Rejected"'},
        {"name": "roles", "label": "Vai trò (phân cách bằng dấu phẩy)", "required": False, "width": 40,
         "desc": "Danh sách vai trò Frappe, phân cách bằng dấu phẩy. "
                 "Chỉ thêm, không xóa vai trò hiện có. "
                 "VD: HTM Technician, HTM Manager.",
         "example": "HTM Technician, HTM Manager"},
    ], [{}],
    instructions="Mỗi hàng = 1 người dùng. "
                 "Email là định danh duy nhất. Nếu đã tồn tại sẽ cập nhật thông tin. "
                 "Vai trò chỉ được thêm, không bao giờ xóa vai trò hiện có.")

    wb.save(os.path.join(OUT_DIR, "06_danh_sach_nguoi_dung.xlsx"))
    print("✓ 06_danh_sach_nguoi_dung.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# 7. HƯỚNG DẪN TỔNG QUAN
# ════════════════════════════════════════════════════════════════════════════

def make_guide():
    wb = Workbook()
    ws = wb.active
    ws.title = "Hướng dẫn Import"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 30

    # Title
    ws.merge_cells("A1:E1")
    c = ws.cell(row=1, column=1,
                value="ASSETCORE — HƯỚNG DẪN IMPORT DỮ LIỆU HÀ LOẠT")
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = fill("1A5276")
    c.alignment = center()
    c.border = BORDER
    ws.row_dimensions[1].height = 36

    headers = ["STT", "File / Sheet", "Mục đích", "Thứ tự import", "Lưu ý quan trọng"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = header_font()
        c.fill = fill(C_HEADER_OPT)
        c.alignment = center()
        c.border = BORDER
    ws.row_dimensions[2].height = 22

    rows = [
        (1, "01_du_lieu_tham_chieu.xlsx\n  → Danh mục tài sản (Category)",
         "Danh mục phân loại thiết bị y tế\n(máy siêu âm, máy thở, ...)",
         "1 — Import đầu tiên",
         "Phải có trước khi tạo Device Model và Asset"),
        (2, "01_du_lieu_tham_chieu.xlsx\n  → Khoa phòng (Department)",
         "Cơ cấu tổ chức bệnh viện / phòng ban",
         "1 — Import đầu tiên",
         "Cấu trúc cây: tạo node cha trước node con"),
        (3, "01_du_lieu_tham_chieu.xlsx\n  → Vị trí (Location)",
         "Vị trí vật lý đặt thiết bị trong bệnh viện",
         "1 — Import đầu tiên",
         "Cấu trúc cây: tạo node cha trước node con"),
        (4, "02_imm00_ncc_model_hopdong_sla.xlsx\n  → Nhà cung cấp",
         "Danh sách NCC / nhà phân phối / phòng HC",
         "2 — Sau bước 1",
         "Cần có trước khi tạo hợp đồng"),
        (5, "02_imm00_ncc_model_hopdong_sla.xlsx\n  → Mô hình thiết bị",
         "Catalog model thiết bị y tế theo hãng",
         "2 — Sau bước 1",
         "Cần danh mục tài sản (Category) đã tồn tại"),
        (6, "02_imm00_ncc_model_hopdong_sla.xlsx\n  → Hợp đồng",
         "Hợp đồng bảo trì / hiệu chuẩn / dịch vụ",
         "3 — Sau NCC",
         "NCC phải tồn tại trước"),
        (7, "02_imm00_ncc_model_hopdong_sla.xlsx\n  → Chính sách SLA",
         "Cam kết mức dịch vụ theo độ ưu tiên",
         "3 — Sau User L1/L2",
         "Escalation user phải tồn tại trước"),
        (8, "06_danh_sach_nguoi_dung.xlsx",
         "Tài khoản người dùng và phân quyền",
         "2 — Sau bước 1",
         "Role Profile phải cấu hình trước"),
        (9, "04_danh_sach_phu_tung.xlsx",
         "Danh mục phụ tùng / vật tư thay thế",
         "3 — Sau NCC",
         "NCC ưu tiên phải tồn tại trước"),
        (10, "05_kho_hang.xlsx",
         "Kho lưu trữ phụ tùng và vật tư",
         "3 — Sau Location & Department",
         "Vị trí và khoa phòng phải tồn tại"),
        (11, "03_danh_sach_tai_san.xlsx",
         "Toàn bộ tài sản thiết bị y tế",
         "4 — CUỐI CÙNG",
         "Cần: Category, Model, NCC, Location, Department, User"),
    ]

    alt_colors = ["FDFEFE", "EAF2FF"]
    for i, row in enumerate(rows):
        r = i + 3
        ws.row_dimensions[r].height = 50
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill(alt_colors[i % 2])
            c.border = BORDER
            c.alignment = Alignment(horizontal="left" if col > 1 else "center",
                                    vertical="center", wrap_text=True)
            c.font = cell_font(size=10, bold=(col == 4))

    # Legend
    ws.row_dimensions[len(rows) + 4].height = 18
    legend_row = len(rows) + 4
    ws.merge_cells(f"A{legend_row}:E{legend_row}")
    c = ws.cell(row=legend_row, column=1,
                value="QUY ƯỚC MÀU:  ■ Đỏ đậm = cột bắt buộc   ■ Xanh đậm = cột tùy chọn   "
                      "■ Xanh nhạt = hàng mô tả   ■ Xanh lá = hàng ví dụ")
    c.font = Font(name="Calibri", bold=True, size=10, color="1A237E")
    c.fill = fill("EDE7F6")
    c.alignment = left()
    c.border = BORDER

    ws.freeze_panes = "A3"
    wb.save(os.path.join(OUT_DIR, "00_huong_dan_import.xlsx"))
    print("✓ 00_huong_dan_import.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print(f"\nGenerating AssetCore import templates → {OUT_DIR}\n")
    make_guide()
    make_asset_category()
    make_department()
    make_location()
    make_imm00()
    make_assets()
    make_spare_parts()
    make_warehouse()
    make_users()
    print("\nDone! 9 files generated.")
