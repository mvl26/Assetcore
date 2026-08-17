# assetcore/tests/guards/test_import_enum_labels.py
# Copyright (c) 2026, AssetCore Team
"""Guard — cột Select trong file nhập/xuất phải hỏi bằng TIẾNG VIỆT (LL-IMP-7).

Bài học 2026-08-11: mẫu bảng kiểm được Việt hoá enum trước, 7 mẫu còn lại vẫn bắt
người dùng gõ "Semi-Annual" / "Pass/Fail" / "Straight Line". Nửa Việt nửa Anh còn
khó dùng hơn thuần Anh vì người ta không đoán được cột nào theo luật nào.

Bốn tầng khoá, đủ để một cột Select MỚI không thể lọt ra ngoài:
  1. PHỦ KÍN  — mọi cột Select có trong file mẫu hoặc file xuất phải có nhãn VI
                cho ĐỦ mọi lựa chọn của DocType.
  2. KHÔNG RÁC — mọi khoá nhãn phải là lựa chọn CÓ THẬT (chống gõ sai enum).
  3. PARITY FE — nhãn phải TRÙNG SSoT của giao diện; lệch = người dùng đọc màn
                hình một kiểu, điền file một kiểu.
  4. FILE THẬT — dropdown + ô ví dụ trong .xlsx phải là nhãn VI, không phải enum.
"""
from __future__ import annotations

import re
from pathlib import Path

import frappe
from frappe.tests.utils import FrappeTestCase
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from assetcore.utils.import_helpers import (
    ENUM_DISPLAY_BY_DOCTYPE,
    SUPPORTED_REF_DOCTYPES,
    _REF_DATA_CONFIG,
    _SHEET_NAME_MAP,
    _TEMPLATE_MAP,
    enum_accepted,
    enum_display,
    enum_to_stored,
    get_template_path,
)
from assetcore.tests._helpers.paths import FRONTEND_SRC

_FE_SRC = Path(FRONTEND_SRC)

# Cột Select KHÔNG cần nhãn VI — lý do phải nêu rõ, không được thêm bừa.
_EXEMPT: set[tuple[str, str]] = {
    # naming_series là mã kỹ thuật do hệ thống sinh, không bao giờ cho người dùng chọn.
}


def _template_columns(doctype: str) -> list[str]:
    if doctype not in _TEMPLATE_MAP:
        return []
    wb = load_workbook(get_template_path(doctype), data_only=True)
    sheet = _SHEET_NAME_MAP.get(doctype, "")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    return [str(c.value).strip() for c in ws[2] if c.value]


def _select_options(doctype: str, fieldname: str) -> list[str] | None:
    """Lựa chọn thật của một cột Select; None nếu không phải Select."""
    try:
        field = frappe.get_meta(doctype).get_field(fieldname)
    except Exception:
        return None
    if not field or field.fieldtype != "Select":
        return None
    return [o.strip() for o in (field.options or "").split("\n") if o.strip()]


def _user_facing_columns(doctype: str) -> set[str]:
    """Cột người dùng THẤY: có trong file mẫu hoặc trong file xuất."""
    cols = set(_template_columns(doctype))
    cfg = _REF_DATA_CONFIG.get(doctype, {})
    cols |= set(cfg.get("export_fields", []))
    cols |= set(cfg.get("child_fields", []))
    return cols


def _parse_ts_map(path: Path, const_name: str) -> dict[str, str]:
    """Đọc một map nhãn từ mã nguồn FE (SSoT hiển thị) — chỉ cặp chuỗi phẳng."""
    source = path.read_text(encoding="utf-8")
    match = re.search(
        rf"const\s+{const_name}\s*:\s*Record<string,\s*string>\s*=\s*\{{(.*?)\n\}}",
        source, re.S,
    )
    assert match, f"không tìm thấy {const_name} trong {path.name}"
    out: dict[str, str] = {}
    for line in match.group(1).splitlines():
        line = line.split("//")[0].strip()
        m = re.match(r"^'([^']+)'\s*:\s*'([^']*)'|^([A-Za-z_][\w]*)\s*:\s*'([^']*)'", line)
        if m:
            key = m.group(1) or m.group(3)
            out[key] = m.group(2) if m.group(1) else m.group(4)
    return out


class TestImportEnumCoverage(FrappeTestCase):
    """Tầng 1+2 — phủ kín và không rác."""

    def test_every_user_facing_select_has_vietnamese_labels(self):
        missing: list[str] = []
        for doctype in set(SUPPORTED_REF_DOCTYPES) | set(_TEMPLATE_MAP):
            if not frappe.db.exists("DocType", doctype):
                continue
            declared = ENUM_DISPLAY_BY_DOCTYPE.get(doctype, {})
            for column in sorted(_user_facing_columns(doctype)):
                if (doctype, column) in _EXEMPT:
                    continue
                options = _select_options(doctype, column)
                if not options or len(options) < 2:
                    continue
                labels = declared.get(column)
                if labels is None:
                    missing.append(f"{doctype}.{column} (thiếu hẳn nhãn VI)")
                    continue
                gaps = [o for o in options if o not in labels]
                if gaps:
                    missing.append(f"{doctype}.{column} thiếu nhãn cho: {gaps}")
        self.assertEqual(
            missing, [],
            "Cột Select người dùng thấy mà chưa có nhãn tiếng Việt ⇒ file mẫu/"
            f"file xuất lại bắt gõ enum tiếng Anh: {missing}",
        )

    def test_no_label_for_a_choice_that_does_not_exist(self):
        """Nhãn cho lựa chọn không có thật = gõ sai enum, im lặng vô hiệu."""
        stray: list[str] = []
        for doctype, fields in ENUM_DISPLAY_BY_DOCTYPE.items():
            if not frappe.db.exists("DocType", doctype):
                continue
            for column, labels in fields.items():
                options = _select_options(doctype, column)
                if options is None:
                    # Bảng con: cột nằm trên DocType con, tra ở đó.
                    child = _REF_DATA_CONFIG.get(doctype, {}).get("child_doctype")
                    options = _select_options(child, column) if child else None
                if options is None:
                    continue
                extra = [k for k in labels if k not in options]
                if extra:
                    stray.append(f"{doctype}.{column}: {extra} không nằm trong {options}")
        self.assertEqual(stray, [], f"Khoá nhãn không khớp lựa chọn thật: {stray}")

    def test_labels_are_unique_within_a_column(self):
        """2 lựa chọn cùng nhãn ⇒ đổi ngược không xác định, dữ liệu vào sai ô."""
        for doctype, fields in ENUM_DISPLAY_BY_DOCTYPE.items():
            for column, labels in fields.items():
                values = list(labels.values())
                self.assertEqual(
                    len(values), len(set(values)),
                    f"{doctype}.{column} có nhãn trùng nhau: {values}",
                )

    def test_accept_both_and_convert_back(self):
        """Nhận CẢ nhãn VI lẫn giá trị gốc; đổi ngược ra đúng giá trị lưu."""
        for doctype, fields in ENUM_DISPLAY_BY_DOCTYPE.items():
            for column, labels in fields.items():
                accepted = enum_accepted(doctype, column)
                for stored, vi in labels.items():
                    self.assertIn(stored, accepted)
                    self.assertIn(vi, accepted)
                    self.assertEqual(enum_to_stored(doctype, column, vi), stored)
                    self.assertEqual(enum_to_stored(doctype, column, stored), stored)
                    self.assertEqual(enum_display(doctype, column, stored), vi)

    def test_unknown_value_passes_through_untouched(self):
        """Giá trị lạ KHÔNG bị bịa — validator lo việc báo lỗi, không phải bộ đổi."""
        self.assertEqual(enum_to_stored("AC Asset", "lifecycle_status", "Xyz"), "Xyz")
        self.assertEqual(enum_display("AC Asset", "lifecycle_status", "Xyz"), "Xyz")
        self.assertEqual(enum_to_stored("AC Asset", "khong_co_cot_nay", "Xyz"), "Xyz")


class TestImportEnumParityWithFrontend(FrappeTestCase):
    """Tầng 3 — nhãn file phải trùng nhãn màn hình."""

    _PAIRS = [
        # (file FE, tên map FE, doctype, cột BE)
        ("constants/labels.ts", "MEDICAL_DEVICE_CLASS_LABEL",
         "IMM Device Model", "medical_device_class"),
        ("constants/labels.ts", "CONTRACT_TYPE_LABEL",
         "Service Contract", "contract_type"),
        ("constants/labels.ts", "MEASUREMENT_TYPE_LABELS",
         "PM Checklist Template", "measurement_type"),
        ("utils/formatters.ts", "DEPRECIATION_METHOD_MAP",
         "AC Asset Category", "default_depreciation_method"),
        ("utils/formatters.ts", "PM_TYPE_MAP", "PM Checklist Template", "pm_type"),
        ("utils/formatters.ts", "FREQUENCY_MAP", "AC Asset", "depreciation_frequency"),
        ("utils/formatters.ts", "STATUS_MAP", "AC Asset", "lifecycle_status"),
        ("utils/formatters.ts", "STATUS_MAP", "User", "imm_approval_status"),
        ("constants/labels.ts", "CLINICAL_AREA_TYPE_LABEL",
         "AC Location", "clinical_area_type"),
        ("constants/labels.ts", "INFECTION_CONTROL_LEVEL_LABEL",
         "AC Location", "infection_control_level"),
        ("constants/labels.ts", "SPARE_PART_CATEGORY_LABEL",
         "AC Spare Part", "part_category"),
        ("constants/labels.ts", "VENDOR_TYPE_LABEL", "AC Supplier", "supplier_group"),
        ("constants/labels.ts", "VENDOR_TYPE_LABEL", "AC Supplier", "vendor_type"),
        ("constants/labels.ts", "INCIDENT_SEVERITY_LABEL",
         "IMM SLA Policy", "risk_class"),
    ]

    def test_backend_labels_match_frontend_ssot(self):
        drift: list[str] = []
        for rel, const_name, doctype, column in self._PAIRS:
            fe_map = _parse_ts_map(_FE_SRC / rel, const_name)
            be_map = ENUM_DISPLAY_BY_DOCTYPE[doctype][column]
            for stored, vi in be_map.items():
                fe_label = fe_map.get(stored)
                if fe_label is None:
                    continue   # FE chưa phủ giá trị này — không phải drift
                if fe_label != vi:
                    drift.append(
                        f"{doctype}.{column}['{stored}']: file='{vi}' ≠ màn hình='{fe_label}' "
                        f"({rel}::{const_name})",
                    )
        self.assertEqual(
            drift, [],
            "Nhãn trong file nhập/xuất lệch nhãn trên giao diện ⇒ người dùng đọc "
            f"màn hình một kiểu, điền file một kiểu: {drift}",
        )


class TestImportTemplatesUseVietnameseChoices(FrappeTestCase):
    """Tầng 4 — kiểm trên FILE .xlsx thật, không tin cấu hình."""

    @staticmethod
    def _sheet(doctype: str):
        wb = load_workbook(get_template_path(doctype), data_only=True)
        sheet = _SHEET_NAME_MAP.get(doctype, "")
        return wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

    def test_dropdown_and_example_use_vietnamese_labels(self):
        offenders: list[str] = []
        for doctype in sorted(set(_TEMPLATE_MAP) & set(ENUM_DISPLAY_BY_DOCTYPE)):
            ws = self._sheet(doctype)
            columns = {
                get_column_letter(c.column): str(c.value).strip()
                for c in ws[2] if c.value
            }
            declared = ENUM_DISPLAY_BY_DOCTYPE[doctype]

            # (a) dropdown phải liệt kê đúng nhãn VI
            for dv in ws.data_validations.dataValidation:
                letter = str(dv.sqref).split(":")[0].rstrip("0123456789")
                column = columns.get(letter)
                if column not in declared:
                    continue
                choices = [c.strip() for c in (dv.formula1 or "").strip('"').split(",")]
                expected = list(declared[column].values())
                if choices != expected:
                    offenders.append(f"{doctype}.{column} dropdown={choices} ≠ {expected}")

            # (b) ô ví dụ (hàng 5) phải là nhãn VI — người dùng copy y hệt
            for letter, column in columns.items():
                if column not in declared:
                    continue
                example = str(ws[f"{letter}5"].value or "").strip()
                if example and example not in declared[column].values():
                    offenders.append(
                        f"{doctype}.{column} ví dụ='{example}' không phải nhãn VI "
                        f"({list(declared[column].values())})",
                    )
        self.assertEqual(offenders, [], f"File mẫu còn hỏi bằng enum tiếng Anh: {offenders}")
