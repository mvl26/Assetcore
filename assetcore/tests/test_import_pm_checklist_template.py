# assetcore/tests/test_import_pm_checklist_template.py
# Copyright (c) 2026, AssetCore Team
"""TDD — Nhập/xuất hàng loạt MẪU BẢNG KIỂM bảo trì (/pm/templates).

Khác mọi loại dữ liệu đang hỗ trợ: đây là CHA + BẢNG CON. File phẳng, mỗi hàng =
1 hạng mục kiểm tra, các cột của mẫu lặp lại; hệ thống gộp theo khoá định danh
(Danh mục tài sản + Loại bảo trì) — đúng `autoname: PMCT-{asset_category}-{pm_type}`.

Acceptance:
  - Nhiều hàng cùng khoá ⇒ MỘT mẫu với N hạng mục, đúng thứ tự hàng.
  - Danh mục điền bằng TÊN; loại bảo trì / cách ghi nhận điền bằng nhãn tiếng
    Việt ("Hàng quý", "Số đo") ⇒ lưu ra giá trị gốc của DocType.
  - Mẫu đã tồn tại (cùng danh mục + loại) ⇒ lỗi gắn vào MỌI hàng của mẫu đó;
    bật 'bỏ qua dòng lỗi/trùng' ⇒ mẫu trùng bị loại, mẫu còn lại vẫn vào.
  - Hàng lỗi lẻ trong một mẫu ⇒ bỏ hàng đó, mẫu vẫn tạo với các hạng mục còn lại.
  - Xuất Excel trải phẳng đúng bố cục file nhập (danh mục = TÊN, enum = nhãn VI).
"""
from __future__ import annotations

import io

import frappe
from frappe.tests.utils import FrappeTestCase

from assetcore.api.import_data import _do_import, _do_preview
from assetcore.services.import_validators import get_validator
from assetcore.utils.import_helpers import (
    GROUPED_IMPORT_DOCTYPES,
    SUPPORTED_REF_DOCTYPES,
    export_ref_data,
    get_template_path,
)

_DOCTYPE = "PM Checklist Template"
_CAT_A = "_TEST Bảng kiểm — Máy thở"
_CAT_B = "_TEST Bảng kiểm — Máy siêu âm"

# Cột đúng thứ tự template 07_bang_kiem_bao_tri.xlsx
_FIELDS = [
    "template_name", "asset_category", "pm_type", "version", "effective_date",
    "description", "measurement_type", "unit",
    "expected_min", "expected_max", "is_critical", "reference_section",
]


def _row(**kw) -> list:
    return [str(kw.get(f, "")) for f in _FIELDS]


class TestImportPmChecklistTemplate(FrappeTestCase):
    _files: list[str] = []

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.cats = {}
        for label in (_CAT_A, _CAT_B):
            doc = frappe.get_doc({
                "doctype": "AC Asset Category", "category_name": label,
            }).insert(ignore_permissions=True)
            cls.cats[label] = doc.name
        frappe.db.commit()

    @classmethod
    def tearDownClass(cls):
        cls._purge()
        for name in cls.cats.values():
            if frappe.db.exists("AC Asset Category", name):
                frappe.delete_doc("AC Asset Category", name,
                                  force=True, ignore_permissions=True)
        for f in cls._files:
            if frappe.db.exists("File", f):
                frappe.delete_doc("File", f, force=True, ignore_permissions=True)
        frappe.db.commit()
        super().tearDownClass()

    @classmethod
    def _purge(cls):
        for cat in cls.cats.values():
            for t in frappe.get_all(_DOCTYPE, filters={"asset_category": cat}):
                frappe.delete_doc(_DOCTYPE, t.name, force=True, ignore_permissions=True)
        frappe.db.commit()

    def tearDown(self):
        self._purge()
        super().tearDown()

    # ── helper: dựng file CSV đúng layout template (5 hàng khung + dữ liệu) ──
    def _save_csv(self, data_rows: list[list]) -> str:
        import csv

        from frappe.utils.file_manager import save_file

        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["BANNER"])
        w.writerow(_FIELDS)
        w.writerow(["nhãn"] * len(_FIELDS))
        w.writerow(["mô tả"] * len(_FIELDS))
        w.writerow(["ví dụ"] * len(_FIELDS))
        for r in data_rows:
            w.writerow(r)

        fdoc = save_file(
            f"pmct_{frappe.generate_hash(length=8)}.csv",
            buf.getvalue().encode("utf-8-sig"), "", "", is_private=1,
        )
        self.__class__._files.append(fdoc.name)
        frappe.db.commit()
        return fdoc.file_url

    # ── wiring ──────────────────────────────────────────────────────────────

    def test_doctype_is_wired_into_the_import_pipeline(self):
        self.assertIn(_DOCTYPE, SUPPORTED_REF_DOCTYPES)
        self.assertIn(_DOCTYPE, GROUPED_IMPORT_DOCTYPES)
        self.assertEqual(get_validator(_DOCTYPE).doctype, _DOCTYPE)
        self.assertTrue(get_template_path(_DOCTYPE).endswith("07_bang_kiem_bao_tri.xlsx"))

    # ── nhập theo nhóm ──────────────────────────────────────────────────────

    def test_rows_of_one_template_become_one_record_with_all_items(self):
        """3 hàng cùng danh mục + loại ⇒ 1 mẫu, 3 hạng mục, đúng thứ tự."""
        file_url = self._save_csv([
            _row(template_name="Bảng kiểm quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Kiểm tra áp suất khí",
                 measurement_type="Số đo", unit="bar",
                 expected_min="4", expected_max="6", is_critical="1"),
            _row(template_name="Bảng kiểm quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Vệ sinh bộ lọc",
                 measurement_type="Đạt/Không đạt"),
            _row(template_name="Bảng kiểm quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Ghi nhận giờ chạy máy",
                 measurement_type="Ghi chú"),
        ])
        res = _do_import(_DOCTYPE, file_url)
        self.assertEqual(res["failed"], 0, res)
        self.assertEqual(res["groups_created"], 1, res)
        self.assertEqual(res["success"], 3, "success đếm theo DÒNG file")

        names = frappe.get_all(_DOCTYPE, filters={"asset_category": self.cats[_CAT_A]})
        self.assertEqual(len(names), 1, "3 hàng cùng khoá phải gộp thành MỘT mẫu")

        doc = frappe.get_doc(_DOCTYPE, names[0].name)
        self.assertEqual(
            [i.description for i in doc.checklist_items],
            ["Kiểm tra áp suất khí", "Vệ sinh bộ lọc", "Ghi nhận giờ chạy máy"],
            "thứ tự hạng mục phải theo thứ tự hàng trong file",
        )
        # Nhãn VI → giá trị gốc của DocType Select
        self.assertEqual(doc.pm_type, "Quarterly")
        self.assertEqual(doc.checklist_items[0].measurement_type, "Numeric")
        self.assertEqual(doc.checklist_items[1].measurement_type, "Pass/Fail")
        self.assertEqual(doc.checklist_items[2].measurement_type, "Text")
        # Danh mục điền bằng TÊN → lưu ra mã
        self.assertEqual(doc.asset_category, self.cats[_CAT_A])
        self.assertEqual(doc.checklist_items[0].expected_min, 4.0)
        self.assertEqual(doc.checklist_items[0].is_critical, 1)
        # item_code do controller tự đánh số — người dùng KHÔNG phải điền
        self.assertEqual(doc.checklist_items[0].item_code, "ITEM-001")

    def test_two_templates_in_one_file(self):
        file_url = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Áp suất", measurement_type="Số đo"),
            _row(template_name="Năm — Siêu âm", asset_category=_CAT_B,
                 pm_type="Hàng năm", description="Đầu dò", measurement_type="Đạt/Không đạt"),
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Bộ lọc", measurement_type="Đạt/Không đạt"),
        ])
        res = _do_import(_DOCTYPE, file_url)
        self.assertEqual(res["groups_created"], 2, res)
        self.assertEqual(res["failed"], 0, res)
        # Hàng cùng khoá nhưng KHÔNG liền nhau vẫn phải về chung một mẫu
        a = frappe.get_all(_DOCTYPE, filters={"asset_category": self.cats[_CAT_A]})
        self.assertEqual(len(a), 1)
        self.assertEqual(len(frappe.get_doc(_DOCTYPE, a[0].name).checklist_items), 2)

    def test_category_accepts_system_code_too(self):
        """Điền mã (từ file xuất cũ) vẫn nhận — cùng khoá với điền tên."""
        file_url = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Hạng mục 1", measurement_type="Ghi chú"),
            _row(template_name="Quý — Máy thở", asset_category=self.cats[_CAT_A],
                 pm_type="Quarterly", description="Hạng mục 2", measurement_type="Ghi chú"),
        ])
        res = _do_import(_DOCTYPE, file_url)
        self.assertEqual(res["groups_created"], 1,
                         "tên và mã của cùng danh mục phải rơi vào MỘT nhóm")

    # ── kiểm tra trước khi nhập ─────────────────────────────────────────────

    def test_preview_reports_row_and_vietnamese_column(self):
        file_url = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="", measurement_type="Số đo"),
        ])
        prev = _do_preview(_DOCTYPE, file_url)
        errs = [e for e in prev["errors"] if e["field"] == "description"]
        self.assertEqual(len(errs), 1, prev["errors"])
        self.assertEqual(errs[0]["source_row"], 6, "hàng dữ liệu đầu tiên = hàng 6")
        self.assertEqual(errs[0]["label"], "Nội dung kiểm tra")
        self.assertEqual(prev["field_labels"]["pm_type"], "Loại bảo trì định kỳ")

    def test_invalid_measurement_and_bounds_are_blocked(self):
        file_url = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Sai cách ghi",
                 measurement_type="Chụp ảnh"),
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Ngưỡng ngược",
                 measurement_type="Số đo", expected_min="9", expected_max="2"),
        ])
        prev = _do_preview(_DOCTYPE, file_url)
        fields = {e["field"] for e in prev["errors"]}
        self.assertIn("measurement_type", fields)
        self.assertIn("expected_max", fields)

    def test_unknown_category_names_the_fix(self):
        file_url = self._save_csv([
            _row(template_name="Quý — Máy lạ", asset_category="Danh mục không có thật",
                 pm_type="Hàng quý", description="X", measurement_type="Ghi chú"),
        ])
        prev = _do_preview(_DOCTYPE, file_url)
        msgs = [e["message"] for e in prev["errors"] if e["field"] == "asset_category"]
        self.assertTrue(msgs, prev["errors"])
        self.assertIn("TÊN", msgs[0], "phải nhắc điền TÊN, không điền mã")

    # ── trùng + bỏ qua ──────────────────────────────────────────────────────

    def test_existing_template_blocks_every_row_of_that_group_only(self):
        """Mẫu trùng ⇒ mọi hàng của mẫu đó lỗi; mẫu khác trong file vẫn nhập được."""
        first = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Hạng mục cũ", measurement_type="Ghi chú"),
        ])
        self.assertEqual(_do_import(_DOCTYPE, first)["groups_created"], 1)

        second = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Hạng mục mới 1", measurement_type="Ghi chú"),
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Hạng mục mới 2", measurement_type="Ghi chú"),
            _row(template_name="Năm — Siêu âm", asset_category=_CAT_B,
                 pm_type="Hàng năm", description="Đầu dò", measurement_type="Đạt/Không đạt"),
        ])
        prev = _do_preview(_DOCTYPE, second)
        dup_rows = {e["row"] for e in prev["errors"]}
        self.assertEqual(dup_rows, {1, 2}, "chỉ 2 hàng của mẫu trùng bị chặn")

        # Chế độ nghiêm ngặt: chặn cả file
        from assetcore.services.shared import ServiceError
        with self.assertRaises(ServiceError):
            _do_import(_DOCTYPE, second, skip_invalid=False)

        # Bỏ qua dòng lỗi/trùng: mẫu trùng bị loại, mẫu mới vẫn vào
        res = _do_import(_DOCTYPE, second, skip_invalid=True)
        self.assertEqual(res["skipped"], 2, res)
        self.assertEqual(res["groups_created"], 1, res)
        self.assertEqual(res["failed"], 0, res)
        self.assertTrue(frappe.db.exists(
            _DOCTYPE, {"asset_category": self.cats[_CAT_B], "pm_type": "Annual"}))
        # Mẫu cũ KHÔNG bị ghi đè
        old = frappe.get_all(_DOCTYPE, filters={"asset_category": self.cats[_CAT_A]})
        self.assertEqual(len(old), 1)
        self.assertEqual(
            [i.description for i in frappe.get_doc(_DOCTYPE, old[0].name).checklist_items],
            ["Hạng mục cũ"],
        )

    def test_bad_row_is_skipped_but_the_template_still_gets_its_good_items(self):
        file_url = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Hạng mục tốt 1",
                 measurement_type="Ghi chú"),
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="",           # thiếu nội dung
                 measurement_type="Ghi chú"),
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Hạng mục tốt 2",
                 measurement_type="Ghi chú"),
        ])
        res = _do_import(_DOCTYPE, file_url, skip_invalid=True)
        self.assertEqual(res["skipped"], 1, res)
        self.assertEqual(res["groups_created"], 1, res)

        name = frappe.get_all(_DOCTYPE, filters={"asset_category": self.cats[_CAT_A]})[0].name
        self.assertEqual(
            [i.description for i in frappe.get_doc(_DOCTYPE, name).checklist_items],
            ["Hạng mục tốt 1", "Hạng mục tốt 2"],
            "hàng lỗi bị bỏ, phần còn lại của mẫu vẫn vào",
        )

    def test_conflicting_parent_columns_warn_and_first_row_wins(self):
        file_url = self._save_csv([
            _row(template_name="Tên A", asset_category=_CAT_A, pm_type="Hàng quý",
                 version="1.0", description="H1", measurement_type="Ghi chú"),
            _row(template_name="Tên B", asset_category=_CAT_A, pm_type="Hàng quý",
                 version="2.0", description="H2", measurement_type="Ghi chú"),
        ])
        prev = _do_preview(_DOCTYPE, file_url)
        warn_fields = {w["field"] for w in prev["warnings"]}
        self.assertIn("template_name", warn_fields)
        self.assertIn("version", warn_fields)
        self.assertEqual(prev["errors"], [], "khai lệch chỉ là cảnh báo, không chặn")

        _do_import(_DOCTYPE, file_url)
        name = frappe.get_all(_DOCTYPE, filters={"asset_category": self.cats[_CAT_A]})[0].name
        doc = frappe.get_doc(_DOCTYPE, name)
        self.assertEqual(doc.template_name, "Tên A", "hàng đầu nhóm thắng")
        self.assertEqual(doc.version, "1.0")

    # ── xuất ────────────────────────────────────────────────────────────────

    def test_export_flattens_and_uses_display_values(self):
        from openpyxl import load_workbook

        file_url = self._save_csv([
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Áp suất",
                 measurement_type="Số đo", unit="bar"),
            _row(template_name="Quý — Máy thở", asset_category=_CAT_A,
                 pm_type="Hàng quý", description="Bộ lọc",
                 measurement_type="Đạt/Không đạt"),
        ])
        _do_import(_DOCTYPE, file_url)

        ws = load_workbook(io.BytesIO(export_ref_data(_DOCTYPE)), data_only=True).active
        fieldnames = [str(c.value or "") for c in ws[2]]
        rows = [
            dict(zip(fieldnames, r))
            for r in ws.iter_rows(min_row=3, values_only=True)
        ]
        mine = [r for r in rows if r.get("asset_category") == _CAT_A]
        self.assertEqual(len(mine), 2, "mỗi hạng mục = 1 hàng, cột cha lặp lại")
        self.assertEqual(ws.cell(row=1, column=3).value, "Danh mục tài sản")
        self.assertEqual({r["template_name"] for r in mine}, {"Quý — Máy thở"})
        self.assertEqual({r["pm_type"] for r in mine}, {"Hàng quý"},
                         "loại bảo trì xuất ra nhãn tiếng Việt")
        self.assertEqual(
            {r["measurement_type"] for r in mine}, {"Số đo", "Đạt/Không đạt"},
        )
        self.assertNotIn(self.cats[_CAT_A], {r["asset_category"] for r in mine},
                         "cột danh mục KHÔNG được in mã hệ thống")
