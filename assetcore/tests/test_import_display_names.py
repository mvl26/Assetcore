# assetcore/tests/test_import_display_names.py
# Copyright (c) 2026, AssetCore Team
"""TDD — Import/Export điền TÊN thay vì MÃ + báo lỗi chỉ đúng hàng/cột.

Acceptance:
  - Mọi cột Link trong template/export đều nằm ở SSoT `LINK_DISPLAY_BY_DOCTYPE`
    (trừ Link tới `User` — PK là email, chính là giá trị hiển thị). Thiếu 1 entry
    = người dùng buộc phải tra mã hệ thống.
  - Template KHÔNG bảo người dùng điền mã ở cột tham chiếu: mô tả phải nói "TÊN",
    ví dụ mẫu phải là tên thật (`Khoa Hồi sức tích cực`) chứ không phải `Khoa-HSTC`.
  - Export in TÊN hiển thị cho cột Link (round-trip: file xuất ra nhập lại được).
  - Lỗi mang `source_row` = số hàng THẬT trong file (dòng trống KHÔNG làm lệch số)
    và `label` = nhãn tiếng Việt của cột.
  - Chế độ 'bỏ qua dòng lỗi/trùng' áp dụng cho MỌI doctype hỗ trợ, kể cả `User`.
"""
from __future__ import annotations

import io
from pathlib import Path

import frappe
from frappe.tests.utils import FrappeTestCase

from assetcore.api.import_data import (
    _OPTIONAL_LINKS_BY_DOCTYPE,
    _RESOLVABLE_LINKS_BY_DOCTYPE,
    _do_import_users,
    _normalise_row,
)
from assetcore.utils.import_helpers import (
    FIRST_DATA_ROW,
    LINK_DISPLAY_BY_DOCTYPE,
    SOURCE_ROW_KEY,
    SUPPORTED_REF_DOCTYPES,
    _REF_DATA_CONFIG,
    _TEMPLATE_MAP,
    _rows_to_dicts,
    build_error_report,
    enrich_issues,
    export_ref_data,
    field_label,
    get_template_path,
    source_row_of,
)

_APP_ROOT = Path(__file__).resolve().parents[1]

# Link fields trỏ tới User: khoá chính = email = giá trị người dùng gõ, nên
# KHÔNG cần (và không được) khai vào map display→code.
_USER_TARGET = "User"


def _link_fields(doctype: str) -> list[tuple[str, str]]:
    """(fieldname, target_doctype) của mọi Link field thật trên DocType."""
    meta = frappe.get_meta(doctype)
    return [
        (f.fieldname, f.options)
        for f in meta.fields
        if f.fieldtype == "Link" and f.options
    ]


class TestImportLinkDisplaySSoT(FrappeTestCase):
    """Map display→code phải phủ hết cột Link mà người dùng nhập/nhìn thấy."""

    def test_every_exported_link_column_is_in_ssot(self):
        """Cột Link xuất ra Excel mà không có trong SSoT ⇒ file export in MÃ.

        Đây là RED-guard cho đúng lỗi user báo: "phải điền bằng mã". Thêm 1 cột
        Link vào `export_fields` mà quên khai map ⇒ test này đỏ.
        """
        missing: list[str] = []
        for doctype in SUPPORTED_REF_DOCTYPES:
            if not frappe.db.exists("DocType", doctype):
                continue
            links = dict(_link_fields(doctype))
            declared = LINK_DISPLAY_BY_DOCTYPE.get(doctype, {})
            for field in _REF_DATA_CONFIG[doctype]["export_fields"]:
                target = links.get(field)
                if not target or target == _USER_TARGET:
                    continue
                if field not in declared:
                    missing.append(f"{doctype}.{field} → {target}")
        self.assertEqual(
            missing, [],
            "Cột Link thiếu entry trong LINK_DISPLAY_BY_DOCTYPE ⇒ export in mã "
            f"hệ thống và import bắt gõ mã: {missing}",
        )

    def test_ssot_entries_point_at_real_display_fields(self):
        """(link_doctype, display_field) phải là field CÓ THẬT, nếu không resolver
        im lặng không tìm ra gì và mọi giá trị người dùng gõ đều bị coi là sai."""
        for doctype, mapping in LINK_DISPLAY_BY_DOCTYPE.items():
            for field, (link_dt, display_field) in mapping.items():
                if not frappe.db.exists("DocType", link_dt):
                    continue
                fieldnames = {f.fieldname for f in frappe.get_meta(link_dt).fields}
                self.assertIn(
                    display_field, fieldnames,
                    f"{doctype}.{field} trỏ display_field '{display_field}' "
                    f"không tồn tại trên {link_dt}",
                )

    def test_api_alias_is_the_same_object_as_ssot(self):
        """`api.import_data._RESOLVABLE_LINKS_BY_DOCTYPE` chỉ là alias — KHÔNG
        được fork thành bản sao thứ 2 (sửa 1 nơi phải ăn cả pipeline)."""
        self.assertIs(_RESOLVABLE_LINKS_BY_DOCTYPE, LINK_DISPLAY_BY_DOCTYPE)

    def test_user_email_links_are_optional_not_fatal(self):
        """Email người phụ trách sai chính tả chỉ được làm trống ô đó, KHÔNG
        được giết cả dòng bằng lỗi Frappe tiếng Anh."""
        for doctype, field in (
            ("AC Department", "dept_head"),
            ("AC Location", "dept_head"),
            ("AC Asset", "custodian"),
            ("AC Asset", "responsible_technician"),
        ):
            self.assertEqual(
                _OPTIONAL_LINKS_BY_DOCTYPE.get(doctype, {}).get(field), "User",
                f"{doctype}.{field} phải nằm trong _OPTIONAL_LINKS_BY_DOCTYPE",
            )


class TestImportTemplatesAskForNames(FrappeTestCase):
    """Template Excel phải hỏi TÊN, không hỏi mã."""

    @staticmethod
    def _sheet_rows(doctype: str):
        from openpyxl import load_workbook

        from assetcore.utils.import_helpers import _SHEET_NAME_MAP

        wb = load_workbook(get_template_path(doctype), data_only=True)
        sheet = _SHEET_NAME_MAP.get(doctype, "")
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        return rows[1], rows[3], rows[4]   # fieldnames, desc, example

    def test_link_columns_tell_user_to_fill_name(self):
        """Mỗi cột Link trong template phải có chữ 'TÊN' trong dòng mô tả."""
        offenders: list[str] = []
        for doctype, mapping in LINK_DISPLAY_BY_DOCTYPE.items():
            if doctype not in _TEMPLATE_MAP:
                continue
            fieldnames, descs, _ = self._sheet_rows(doctype)
            for i, fn in enumerate(fieldnames):
                if fn not in mapping:
                    continue
                desc = str(descs[i] or "")
                if "TÊN" not in desc.upper():
                    offenders.append(f"{doctype}.{fn}: {desc!r}")
        self.assertEqual(
            offenders, [],
            f"Cột tham chiếu phải yêu cầu điền TÊN, không điền mã: {offenders}",
        )

    def test_user_department_example_is_a_real_name_not_a_code(self):
        """Ví dụ mẫu cột Khoa/Phòng của template người dùng từng là 'Khoa-HSTC'
        (dạng mã) — người dùng copy y hệt sẽ nhập sai."""
        fieldnames, descs, example = self._sheet_rows("User")
        idx = list(fieldnames).index("ac_department")
        self.assertNotIn("Mã", str(descs[idx] or ""),
                         "mô tả không được yêu cầu 'Mã khoa/phòng'")
        value = str(example[idx] or "")
        self.assertNotIn("-", value,
                         f"ví dụ mẫu phải là tên khoa thật, không phải mã: {value!r}")
        self.assertTrue(value.strip(), "ví dụ mẫu không được rỗng")

    def test_banner_points_at_the_first_real_data_row(self):
        """Banner từng ghi 'từ hàng 5' trong khi parser bắt đầu ở hàng 6 ⇒ dòng
        đầu tiên người dùng gõ bị nuốt im lặng."""
        from openpyxl import load_workbook

        for doctype in ("AC Asset", "User", "AC Department"):
            wb = load_workbook(get_template_path(doctype), data_only=True)
            from assetcore.utils.import_helpers import _SHEET_NAME_MAP
            sheet = _SHEET_NAME_MAP.get(doctype, "")
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            banner = str(ws["A1"].value or "")
            self.assertIn(
                f"HÀNG {FIRST_DATA_ROW}", banner.upper(),
                f"banner template {doctype} phải chỉ đúng hàng dữ liệu đầu tiên: {banner!r}",
            )


class TestImportRowNumbering(FrappeTestCase):
    """Báo lỗi phải chỉ đúng số hàng người dùng nhìn thấy trong Excel."""

    def test_blank_rows_do_not_shift_reported_row(self):
        fieldnames = ["category_name", "gmdn_code"]
        raw = [
            ("Máy thở", "12345"),          # hàng 6
            ("", ""),                       # hàng 7 — trống, bị loại
            ("Máy siêu âm", "67890"),       # hàng 8
        ]
        rows = _rows_to_dicts(fieldnames, raw)
        self.assertEqual(len(rows), 2, "dòng trống phải bị loại khỏi dữ liệu")
        self.assertEqual(rows[0][SOURCE_ROW_KEY], 6)
        self.assertEqual(rows[1][SOURCE_ROW_KEY], 8)
        # Dòng dữ liệu thứ 2 nằm ở hàng 8 — KHÔNG phải hàng 7 (6+1).
        self.assertEqual(source_row_of(rows, 2), 8)

    def test_source_row_falls_back_when_not_recorded(self):
        rows = [{"category_name": "X"}]
        self.assertEqual(source_row_of(rows, 1), FIRST_DATA_ROW)

    def test_parser_key_never_reaches_the_doctype(self):
        """`__source_row__` là khoá kỹ thuật — lọt vào doc.update() sẽ ghi đè
        field lạ / nổ lúc insert."""
        clean = _normalise_row(
            {"category_name": "Máy thở", SOURCE_ROW_KEY: 9}, set(),
        )
        self.assertNotIn(SOURCE_ROW_KEY, clean)
        self.assertEqual(clean["category_name"], "Máy thở")

    def test_enrich_issues_adds_label_and_file_row(self):
        rows = _rows_to_dicts(
            ["asset_category"], [("Máy thở",), ("",), ("Máy X-quang",)],
        )
        issues = [
            {"row": 2, "field": "asset_category", "message": "không tồn tại",
             "severity": "error"},
        ]
        enrich_issues("AC Asset", rows, issues)
        self.assertEqual(issues[0]["label"], "Danh mục tài sản",
                         "phải là nhãn tiếng Việt, không phải fieldname")
        self.assertEqual(issues[0]["source_row"], 8,
                         "phải là hàng thật trong file (dòng trống ở hàng 7)")

    def test_field_label_falls_back_to_fieldname(self):
        self.assertEqual(field_label("AC Asset", "asset_category"), "Danh mục tài sản")
        self.assertEqual(field_label("AC Asset", "cot_la"), "cot_la")
        self.assertEqual(field_label("AC Asset", ""), "")

    def test_error_report_leads_with_the_file_row(self):
        from openpyxl import load_workbook

        rows = _rows_to_dicts(
            ["category_name"], [("Máy thở",), ("",), ("Máy siêu âm",)],
        )
        errors = enrich_issues("AC Asset Category", rows, [
            {"row": 2, "field": "category_name", "message": "đã tồn tại",
             "severity": "error"},
        ])
        wb = load_workbook(io.BytesIO(
            build_error_report(["category_name"], rows, errors, "AC Asset Category"),
        ))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "Hàng trong file")
        self.assertEqual(ws.cell(row=3, column=1).value, 8,
                         "dòng dữ liệu thứ 2 nằm ở hàng 8 của file gốc")
        note = str(ws.cell(row=3, column=4).value or "")
        self.assertIn("Tên danh mục", note,
                      "ghi chú lỗi phải gọi cột bằng nhãn tiếng Việt")


class TestExportRendersDisplayNames(FrappeTestCase):
    """File 'Xuất Excel' phải đọc được: cột Link in TÊN, không in mã."""

    _CAT = "_TEST Danh mục hiển thị"
    _MODEL = "_TEST Model hiển thị"

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.cat = frappe.get_doc({
            "doctype": "AC Asset Category", "category_name": cls._CAT,
        }).insert(ignore_permissions=True)
        cls.model = frappe.get_doc({
            "doctype": "IMM Device Model",
            "model_name": cls._MODEL,
            "manufacturer": "_TEST NSX",
            "asset_category": cls.cat.name,
            "medical_device_class": "Class II",
        }).insert(ignore_permissions=True)
        frappe.db.commit()

    @classmethod
    def tearDownClass(cls):
        frappe.delete_doc("IMM Device Model", cls.model.name,
                          force=True, ignore_permissions=True)
        frappe.delete_doc("AC Asset Category", cls.cat.name,
                          force=True, ignore_permissions=True)
        frappe.db.commit()
        super().tearDownClass()

    def _export_row(self, doctype: str, key_field: str, key_value: str) -> dict:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(export_ref_data(doctype)), data_only=True)
        ws = wb.active
        fieldnames = [str(c.value or "") for c in ws[2]]
        for row in ws.iter_rows(min_row=3, values_only=True):
            record = dict(zip(fieldnames, row))
            if str(record.get(key_field) or "") == key_value:
                return record
        self.fail(f"không tìm thấy {key_value} trong export {doctype}")

    def test_link_column_shows_name_not_system_code(self):
        record = self._export_row("IMM Device Model", "model_name", self._MODEL)
        self.assertEqual(
            record["asset_category"], self._CAT,
            "cột 'Danh mục tài sản' phải in TÊN danh mục",
        )
        self.assertNotEqual(
            record["asset_category"], self.cat.name,
            "KHÔNG được in mã hệ thống (người dùng không tra được mã)",
        )

    def test_system_code_column_still_exported(self):
        """Cột 'Mã hệ thống' (`name`) vẫn giữ — dùng để đối chiếu khi hỗ trợ kỹ
        thuật; chỉ các cột THAM CHIẾU mới đổi sang tên."""
        record = self._export_row("IMM Device Model", "model_name", self._MODEL)
        self.assertEqual(record["name"], self.model.name)


class TestSkipInvalidCoversUsers(FrappeTestCase):
    """'Bỏ qua dòng lỗi/trùng' phải dùng được cho mọi loại dữ liệu."""

    def test_user_import_honours_skipped_rows(self):
        rows = [
            {"email": "khong-phai-email", "first_name": "Sai"},
            {"email": "_test_skip_user@assetcore.test", "first_name": "Hợp lệ"},
        ]
        skipped = [{
            "row": 1, "field": "email", "reason": "pre_validate",
            "message": "Email 'khong-phai-email' không đúng định dạng",
        }]
        try:
            res = _do_import_users(rows, {1}, skipped)
            self.assertEqual(res["success"], 1, "dòng hợp lệ vẫn phải được nhập")
            self.assertEqual(res["failed"], 0, "dòng lỗi bị bỏ qua, không tính là failed")
            self.assertEqual(res["skipped"], 1)
            self.assertEqual(len(res["skipped_rows"]), 1)
            self.assertFalse(frappe.db.exists("User", "khong-phai-email"))
        finally:
            if frappe.db.exists("User", "_test_skip_user@assetcore.test"):
                frappe.delete_doc("User", "_test_skip_user@assetcore.test",
                                  force=True, ignore_permissions=True)
            frappe.db.commit()

    def test_strict_mode_message_names_the_skip_control(self):
        """Câu chặn ở chế độ nghiêm ngặt phải chỉ đúng tên nút người dùng thấy."""
        source = (_APP_ROOT / "api" / "import_data.py").read_text(encoding="utf-8")
        self.assertIn("Bỏ qua dòng lỗi/trùng", source)
        self.assertNotIn(
            "chưa hỗ trợ cho import Người dùng", source,
            "chế độ bỏ qua đã hỗ trợ User — bỏ chặn cũ",
        )
