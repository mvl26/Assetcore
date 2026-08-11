# assetcore/tests/test_import_file_layout.py
# Copyright (c) 2026, AssetCore Team
"""TDD — BỐ CỤC FILE: hàng nào là khung, hàng nào là dữ liệu.

Bối cảnh lỗi (đo 2026-08-11): `export_ref_data` ghi khung 2 hàng (nhãn +
fieldname) rồi đổ dữ liệu từ hàng 3, trong khi parser luôn bỏ đúng 5 hàng đầu.
Hệ quả: người dùng bấm "Xuất Excel" → sửa → "Nhập lại" thì **3 bản ghi đầu tiên
biến mất IM LẶNG** — không lỗi, không cảnh báo, chỉ thiếu dữ liệu.

Acceptance:
  - File xuất ra nhập lại được KHÔNG mất một hàng nào (mọi loại dữ liệu).
  - File xuất theo bố cục CŨ (2 hàng khung, người dùng đã tải về trước bản vá)
    vẫn nhập đủ — không được im lặng nuốt 3 hàng đầu.
  - File mẫu (5 hàng khung, có banner) vẫn bỏ đúng hàng ví dụ.
  - Số hàng báo lỗi (`__source_row__`) là số hàng THẬT của từng bố cục.
"""
from __future__ import annotations

import io

import frappe
from frappe.tests.utils import FrappeTestCase

from assetcore.utils.import_helpers import (
    SOURCE_ROW_KEY,
    export_ref_data,
    parse_upload_file,
)

_CAT_DOCTYPE = "AC Asset Category"
_PREFIX = "_TEST Bố cục"


def _save_xlsx_rows(header_rows: list[list], data_rows: list[list]) -> str:
    """Ghi file .xlsx thô theo đúng số hàng khung yêu cầu, trả file_url."""
    from openpyxl import Workbook
    from frappe.utils.file_manager import save_file

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in header_rows + data_rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)

    fdoc = save_file(
        f"layout_{frappe.generate_hash(length=8)}.xlsx",
        buf.getvalue(), "", "", is_private=1,
    )
    frappe.db.commit()
    return fdoc.file_url


class TestImportFileLayout(FrappeTestCase):
    """Parser phải đọc đúng dữ liệu của CẢ file mẫu lẫn file xuất ra."""

    _files: list[str] = []

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.cats = []
        for i in range(1, 5):
            doc = frappe.get_doc({
                "doctype": _CAT_DOCTYPE, "category_name": f"{_PREFIX} {i}",
            }).insert(ignore_permissions=True)
            cls.cats.append(doc.name)
        frappe.db.commit()

    @classmethod
    def tearDownClass(cls):
        for name in cls.cats:
            if frappe.db.exists(_CAT_DOCTYPE, name):
                frappe.delete_doc(_CAT_DOCTYPE, name, force=True, ignore_permissions=True)
        for f in cls._files:
            if frappe.db.exists("File", f):
                frappe.delete_doc("File", f, force=True, ignore_permissions=True)
        frappe.db.commit()
        super().tearDownClass()

    def _save_bytes(self, content: bytes, ext: str = "xlsx") -> str:
        from frappe.utils.file_manager import save_file

        fdoc = save_file(
            f"layout_{frappe.generate_hash(length=8)}.{ext}", content, "", "",
            is_private=1,
        )
        self.__class__._files.append(fdoc.name)
        frappe.db.commit()
        return fdoc.file_url

    # ── vòng xuất → nhập lại ────────────────────────────────────────────────

    def test_exported_file_reimports_without_losing_any_row(self):
        """Bản ghi trong hệ thống = bản ghi parser đọc lại được từ file xuất ra.

        Bất biến theo SỐ LƯỢNG, không theo vị trí: bố cục cũ nuốt đúng 3 hàng
        ĐẦU nên phép so theo tên bản ghi ở cuối file sẽ xanh giả.
        """
        from openpyxl import load_workbook

        blob = export_ref_data(_CAT_DOCTYPE)
        ws = load_workbook(io.BytesIO(blob), data_only=True).active
        fieldnames = [str(c.value or "") for c in ws[2]]
        self.assertIn("category_name", fieldnames,
                      "hàng 2 vẫn phải là hàng fieldname (hợp đồng với parser)")

        file_url = self._save_bytes(blob)
        _, rows = parse_upload_file(file_url, _CAT_DOCTYPE)

        total = frappe.db.count(_CAT_DOCTYPE)
        self.assertEqual(
            len(rows), total,
            f"xuất {total} danh mục nhưng nhập lại chỉ đọc được {len(rows)} "
            "— hàng đầu file bị nuốt im lặng",
        )

    def test_first_record_of_the_exported_file_survives(self):
        """Bản ghi ĐẦU TIÊN của file xuất ra là nạn nhân số một — phải còn."""
        from openpyxl import load_workbook

        blob = export_ref_data(_CAT_DOCTYPE)
        ws = load_workbook(io.BytesIO(blob), data_only=True).active
        fieldnames = [str(c.value or "") for c in ws[2]]

        first_exported = frappe.get_all(
            _CAT_DOCTYPE, fields=["category_name"], order_by="creation asc", limit=1,
        )[0]["category_name"]

        file_url = self._save_bytes(blob)
        _, rows = parse_upload_file(file_url, _CAT_DOCTYPE)
        self.assertIn("category_name", fieldnames)
        self.assertIn(
            first_exported, {str(r.get("category_name") or "") for r in rows},
            "danh mục đầu tiên biến mất khi nhập lại file vừa xuất",
        )

    # ── tương thích ngược: file xuất theo bố cục CŨ ─────────────────────────

    def test_legacy_two_row_header_file_still_parses_every_row(self):
        """File người dùng đã tải về TRƯỚC bản vá (khung 2 hàng) vẫn đủ dữ liệu."""
        fields = ["name", "category_name", "description"]
        labels = ["Mã hệ thống", "Tên danh mục", "Mô tả"]
        data = [[f"CAT-{i:04d}", f"{_PREFIX} cũ {i}", ""] for i in range(1, 6)]

        file_url = _save_xlsx_rows([labels, fields], data)
        self.__class__._files.append(
            frappe.db.get_value("File", {"file_url": file_url}, "name"))
        _, rows = parse_upload_file(file_url, _CAT_DOCTYPE)

        self.assertEqual(len(rows), 5,
                         "bố cục cũ: 5 bản ghi phải ra 5 dòng, không phải 2")
        self.assertEqual(rows[0]["category_name"], f"{_PREFIX} cũ 1")
        self.assertEqual(rows[0][SOURCE_ROW_KEY], 3,
                         "bản ghi đầu của bố cục cũ nằm ở hàng 3 của file")

    # ── file mẫu: vẫn bỏ đúng hàng ví dụ ────────────────────────────────────

    def test_template_layout_skips_exactly_the_example_row(self):
        fields = ["name", "category_name", "description"]
        header = [
            ["📋 HƯỚNG DẪN IMPORT: Danh mục tài sản  |  Điền từ HÀNG 6"],
            fields,
            ["Mã hệ thống", "Tên danh mục", "Mô tả"],
            ["", "Tên duy nhất của danh mục", ""],
            ["", "Máy chẩn đoán hình ảnh", "VÍ DỤ — hệ thống bỏ qua"],
        ]
        data = [["", f"{_PREFIX} mẫu {i}", ""] for i in range(1, 4)]

        file_url = _save_xlsx_rows(header, data)
        self.__class__._files.append(
            frappe.db.get_value("File", {"file_url": file_url}, "name"))
        _, rows = parse_upload_file(file_url, _CAT_DOCTYPE)

        self.assertEqual(len(rows), 3, "chỉ 3 hàng dữ liệu, hàng ví dụ phải bị bỏ")
        self.assertNotIn(
            "Máy chẩn đoán hình ảnh",
            {r.get("category_name") for r in rows},
            "hàng ví dụ của file mẫu KHÔNG được nhập vào hệ thống",
        )
        self.assertEqual(rows[0][SOURCE_ROW_KEY], 6)

    def test_template_layout_without_banner_is_detected_by_label_row(self):
        """Người dùng xoá hàng banner — vẫn nhận ra là file mẫu nhờ hàng nhãn."""
        fields = ["name", "category_name", "description"]
        header = [
            ["", "", ""],
            fields,
            ["Mã hệ thống", "Tên danh mục", "Mô tả"],
            ["", "Mô tả cột", ""],
            ["", "Máy chẩn đoán hình ảnh", "VÍ DỤ"],
        ]
        data = [["", f"{_PREFIX} không banner", ""]]

        file_url = _save_xlsx_rows(header, data)
        self.__class__._files.append(
            frappe.db.get_value("File", {"file_url": file_url}, "name"))
        _, rows = parse_upload_file(file_url, _CAT_DOCTYPE)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["category_name"], f"{_PREFIX} không banner")
