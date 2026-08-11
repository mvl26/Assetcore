# Copyright (c) 2026, AssetCore Team
"""Import/Export helper utilities — parse uploaded Excel, build error reports, export data."""
from __future__ import annotations

import io
import re
from typing import Any

import frappe

_LABEL_SYSTEM_CODE = "Mã hệ thống"

# Khoá kỹ thuật gắn vào mỗi dòng đã parse: số hàng THẬT trong file người dùng
# (1-based, đúng thanh số dòng của Excel). Dùng để báo lỗi "sai ở dòng nào" —
# KHÔNG được ghi vào DocType, nên mọi consumer phải bỏ qua khoá bắt đầu bằng "__".
SOURCE_ROW_KEY = "__source_row__"

# Hàng đầu tiên chứa dữ liệu người dùng trong template (hàng 1 banner, 2 fieldname,
# 3 nhãn VI, 4 mô tả, 5 ví dụ). Dùng chung cho cả Excel lẫn CSV.
FIRST_DATA_ROW = 6

# Dấu nhận biết file theo bố cục TEMPLATE — banner ở ô A1 do generator ghi ra
# (`docs/res/imports/generate_templates.py::build_sheet`) và do `export_ref_data`
# ghi lại y hệt. Sửa chuỗi này thì phải sửa cả hai nơi (guard: test_import_file_layout).
TEMPLATE_BANNER_PREFIX = "📋 HƯỚNG DẪN IMPORT"

# Bố cục CŨ của file "Xuất Excel" (trước 2026-08-11): chỉ 2 hàng khung — nhãn VI
# rồi fieldname — nên dữ liệu bắt đầu ở hàng 3. File cũ đã nằm trong máy người
# dùng, parser phải còn đọc được: bỏ cứng 5 hàng cho file này = nuốt IM LẶNG 3
# bản ghi đầu (đo 2026-08-11: xuất 78 danh mục, nhập lại còn 75).
LEGACY_EXPORT_FIRST_DATA_ROW = 3

# ─────────────────────────────────────────────────────────────────────────────
# LINK DISPLAY — SSoT (LL-IMP-1 / LL-BE-26)
# ─────────────────────────────────────────────────────────────────────────────
#
# Người dùng LUÔN điền TÊN hiển thị (vd "Khoa Hồi sức tích cực"), không bao giờ
# phải điền mã hệ thống (vd "AC-DEPT-0007"). Map này là NGUỒN DUY NHẤT cho cả 3
# hướng dùng, nên sửa 1 chỗ là cả pipeline đồng bộ:
#   1. import  — resolver `api.import_data._resolve_links` đổi tên → mã trước insert
#   2. import  — validator `_link_lookup_set` chấp nhận CẢ tên lẫn mã
#   3. export  — `export_ref_data` in ra TÊN thay vì mã (round-trip đọc được)
#
# RULE (LL-BE-26): Tree DocType (is_tree=1) BẮT BUỘC khai nsm_parent_field ở đây
# (tự trỏ chính nó) — nếu không, Frappe core nested_set.validate_parent_field nổ
# "Could not find Parent <Doctype>: <display_name>".
#
# Link trỏ tới `User` KHÔNG cần khai: PK của User chính là email = giá trị hiển thị.
LINK_DISPLAY_BY_DOCTYPE: dict[str, dict[str, tuple[str, str]]] = {
    "AC Asset": {
        "asset_category": ("AC Asset Category", "category_name"),
        "device_model":   ("IMM Device Model", "model_name"),
        "location":       ("AC Location", "location_name"),
        "department":     ("AC Department", "department_name"),
        "supplier":       ("AC Supplier", "supplier_name"),
    },
    "AC Location": {
        # Tree DocType — parent self-reference (LL-BE-26)
        "parent_location": ("AC Location", "location_name"),
    },
    "AC Department": {
        # Tree DocType — parent self-reference (LL-BE-26)
        "parent_department": ("AC Department", "department_name"),
    },
    "AC Warehouse": {
        "location":   ("AC Location", "location_name"),
        "department": ("AC Department", "department_name"),
    },
    "AC Spare Part": {
        "preferred_supplier": ("AC Supplier", "supplier_name"),
    },
    "Service Contract": {
        "supplier": ("AC Supplier", "supplier_name"),
    },
    "IMM Device Model": {
        "asset_category": ("AC Asset Category", "category_name"),
    },
    "User": {
        # ac_department: người dùng điền "Khoa Hồi sức tích cực" → AC-DEPT-####
        "ac_department": ("AC Department", "department_name"),
    },
    "PM Checklist Template": {
        "asset_category": ("AC Asset Category", "category_name"),
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# ENUM DISPLAY — nhãn tiếng Việt cho cột Select
# ─────────────────────────────────────────────────────────────────────────────
#
# Cùng lý do với LINK_DISPLAY: người dùng bệnh viện không gõ "Semi-Annual" hay
# "Pass/Fail". Giá trị LƯU giữ nguyên chuỗi DocType Select (đừng đụng enum —
# LL-FE-53); chỉ lớp NHẬP/XUẤT dịch qua lại.
#   - template + export  → in nhãn VI
#   - import             → nhận CẢ nhãn VI lẫn giá trị gốc, đổi về giá trị gốc
#
# Nhãn phải khớp SSoT FE (`utils/formatters.ts::PM_TYPE_MAP`,
# `constants/labels.ts::MEASUREMENT_TYPE_LABELS`) — lệch = người dùng đọc màn
# hình một kiểu, điền file một kiểu.
# Nhãn dùng lại ở nhiều DocType — khai 1 lần để không drift giữa các file mẫu.
_DEPRECIATION_METHOD_VI = {
    "Straight Line": "Đường thẳng",
    "Double Declining": "Số dư giảm dần",
    "Units of Production": "Theo sản lượng",
}
# AC Asset có thêm lựa chọn 'None' = không trích khấu hao (Danh mục thì không).
_ASSET_DEPRECIATION_METHOD_VI = {**_DEPRECIATION_METHOD_VI, "None": "Không khấu hao"}
# formatters.ts::FREQUENCY_MAP (miền *frequency*, KHÁC pm_type — 'Ad-hoc' của
# frequency là 'Theo yêu cầu', của pm_type là 'Đột xuất'; khác biệt có chủ đích).
_DEPRECIATION_FREQUENCY_VI = {
    "Monthly": "Hàng tháng",
    "Quarterly": "Hàng quý",
    "Yearly": "Hàng năm",
}
# ⚠️ `supplier_group` và `vendor_type` KHÔNG cùng tập lựa chọn: cái kết thúc bằng
# "Service Provider", cái kia bằng "Service". Gộp 1 map = sinh khoá rác + 2 nhãn
# trùng nhau trong một cột ⇒ đổi ngược không xác định (guard bắt được 2026-08-11).
_SUPPLIER_GROUP_VI = {
    "Manufacturer": "Nhà sản xuất",
    "Distributor": "Nhà phân phối",
    "Calibration Lab": "Phòng hiệu chuẩn",
    "Service Provider": "Dịch vụ",
}
_VENDOR_TYPE_VI = {
    "Manufacturer": "Nhà sản xuất",
    "Distributor": "Nhà phân phối",
    "Calibration Lab": "Phòng hiệu chuẩn",
    "Service": "Dịch vụ",
}

ENUM_DISPLAY_BY_DOCTYPE: dict[str, dict[str, dict[str, str]]] = {
    "AC Asset Category": {
        "default_depreciation_method": _DEPRECIATION_METHOD_VI,
        "depreciation_frequency": _DEPRECIATION_FREQUENCY_VI,
    },
    "AC Location": {
        "clinical_area_type": {
            "ICU": "Hồi sức tích cực",
            "OR": "Phòng mổ",
            "Lab": "Xét nghiệm",
            "Imaging": "Chẩn đoán hình ảnh",
            "General Ward": "Khoa lâm sàng thường",
            "Storage": "Kho",
            "Office": "Văn phòng",
        },
        "infection_control_level": {
            "Standard": "Tiêu chuẩn",
            "Enhanced": "Tăng cường",
            "Isolation": "Cách ly",
        },
    },
    "AC Supplier": {
        "supplier_group": _SUPPLIER_GROUP_VI,
        "vendor_type": _VENDOR_TYPE_VI,
    },
    "IMM Device Model": {
        # labels.ts::MEDICAL_DEVICE_CLASS_LABEL
        "medical_device_class": {
            "Class I": "Loại I — Rủi ro thấp",
            "Class II": "Loại II — Rủi ro trung bình",
            "Class III": "Loại III — Rủi ro cao",
        },
        "default_calibration_type": {
            "Internal": "Nội bộ",
            "External": "Bên ngoài",
            "Both": "Cả hai",
        },
    },
    "Service Contract": {
        # labels.ts::CONTRACT_TYPE_LABEL
        "contract_type": {
            "Preventive Maintenance": "Bảo trì định kỳ",
            "Calibration": "Hiệu chuẩn",
            "Repair": "Sửa chữa",
            "Full Service": "Toàn diện",
            "Warranty Extension": "Gia hạn bảo hành",
        },
    },
    "User": {
        # formatters.ts::STATUS_MAP
        "imm_approval_status": {
            "Pending": "Chờ xử lý",
            "Approved": "Đã phê duyệt",
            "Rejected": "Bị từ chối",
        },
    },
    "AC Asset": {
        # formatters.ts::STATUS_MAP (vòng đời tài sản)
        "lifecycle_status": {
            "Draft": "Bản nháp",
            "Commissioned": "Đã đưa vào sử dụng",
            "Active": "Đang hoạt động",
            "Under Maintenance": "Đang bảo trì",
            "Under Repair": "Đang sửa chữa",
            "Calibrating": "Đang hiệu chuẩn",
            "Out of Service": "Ngừng hoạt động",
            "Decommissioned": "Đã thanh lý",
        },
        # Cột `status` không có trong file mẫu nhưng CÓ trong file xuất — thiếu
        # nhãn là file xuất lẫn lộn nửa Việt nửa Anh.
        "status": {
            "Submitted": "Đã gửi",
            "Active": "Đang hoạt động",
            "Out of Service": "Ngừng hoạt động",
            "Decommissioned": "Đã thanh lý",
            "Under Repair": "Đang sửa chữa",
            "Calibrating": "Đang hiệu chuẩn",
        },
        "depreciation_method": _ASSET_DEPRECIATION_METHOD_VI,
        "depreciation_frequency": _DEPRECIATION_FREQUENCY_VI,
    },
    "AC Spare Part": {
        "part_category": {
            "Electrical": "Điện",
            "Mechanical": "Cơ khí",
            "Consumable": "Tiêu hao",
            "Filter": "Bộ lọc",
            "Battery": "Pin/Ắc-quy",
            "Sensor": "Cảm biến",
            "Other": "Khác",
        },
    },
    # Sheet "Chính sách SLA" nằm trong file mẫu 02 (chưa nối vào wizard nhập,
    # nhưng người dùng vẫn tải file đó về điền) — dropdown phải hợp lệ + tiếng Việt.
    "IMM SLA Policy": {
        "priority": {
            "P1": "P1 — Khẩn cấp",
            "P2": "P2 — Cao",
            "P3": "P3 — Trung bình",
            "P4": "P4 — Thấp",
        },
        # labels.ts::INCIDENT_SEVERITY_LABEL
        "risk_class": {
            "Low": "Thấp",
            "Medium": "Trung bình",
            "High": "Cao",
            "Critical": "Nghiêm trọng",
        },
    },
    "PM Checklist Template": {
        # formatters.ts::PM_TYPE_MAP
        "pm_type": {
            "Quarterly": "Hàng quý",
            "Semi-Annual": "Nửa năm",
            "Annual": "Hàng năm",
            "Ad-hoc": "Đột xuất",
        },
        # labels.ts::MEASUREMENT_TYPE_LABELS
        "measurement_type": {
            "Pass/Fail": "Đạt/Không đạt",
            "Numeric": "Số đo",
            "Text": "Ghi chú",
        },
    },
}


def enum_display(doctype: str, field: str, value: str) -> str:
    """Giá trị gốc → nhãn VI (dùng khi ghi template/export)."""
    return ENUM_DISPLAY_BY_DOCTYPE.get(doctype, {}).get(field, {}).get(value, value)


def enum_accepted(doctype: str, field: str) -> set[str]:
    """Tập giá trị hợp lệ khi nhập: gồm CẢ giá trị gốc lẫn nhãn VI."""
    mapping = ENUM_DISPLAY_BY_DOCTYPE.get(doctype, {}).get(field, {})
    return set(mapping) | set(mapping.values())


def enum_to_stored(doctype: str, field: str, value: str) -> str:
    """Nhãn VI (hoặc giá trị gốc) → giá trị LƯU vào DocType Select."""
    mapping = ENUM_DISPLAY_BY_DOCTYPE.get(doctype, {}).get(field, {})
    if not mapping or value in mapping:
        return value
    reverse = {vi: en for en, vi in mapping.items()}
    return reverse.get(value, value)

# Mapping doctype → (fieldname, tiếng Việt label, export fields)
_REF_DATA_CONFIG: dict[str, dict] = {
    "AC Asset Category": {
        "name_field": "category_name",
        "export_fields": [
            "name", "category_name", "description", "gmdn_code", "gmdn_term",
            "default_pm_required", "default_pm_interval_days",
            "default_calibration_required", "default_calibration_interval_days",
            "default_depreciation_method", "total_depreciation_months",
            "depreciation_frequency", "default_residual_value_pct",
            "has_radiation", "is_active",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "category_name": "Tên danh mục",
            "description": "Mô tả",
            "gmdn_code": "Mã GMDN",
            "gmdn_term": "Tên GMDN",
            "default_pm_required": "Cần bảo trì ĐK",
            "default_pm_interval_days": "Chu kỳ PM (ngày)",
            "default_calibration_required": "Cần hiệu chuẩn",
            "default_calibration_interval_days": "Chu kỳ HC (ngày)",
            "default_depreciation_method": "PP khấu hao",
            "total_depreciation_months": "Thời gian KH (tháng)",
            "depreciation_frequency": "Tần suất KH",
            "default_residual_value_pct": "Giá trị thu hồi (%)",
            "has_radiation": "Có bức xạ",
            "is_active": "Đang hoạt động",
        },
    },
    "AC Department": {
        "name_field": "department_name",
        "export_fields": [
            "name", "department_name", "department_code", "parent_department",
            "is_group", "dept_head", "phone", "email", "is_active",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "department_name": "Tên khoa/phòng",
            "department_code": "Mã khoa",
            "parent_department": "Khoa cha",
            "is_group": "Là nhóm cha",
            "dept_head": "Trưởng khoa",
            "phone": "Điện thoại",
            "email": "Email",
            "is_active": "Đang hoạt động",
        },
    },
    "AC Location": {
        "name_field": "location_name",
        "export_fields": [
            "name", "location_name", "location_code", "parent_location",
            "is_group", "clinical_area_type", "infection_control_level",
            "power_backup_available", "dept_head", "contact_phone", "notes",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "location_name": "Tên vị trí",
            "location_code": "Mã vị trí",
            "parent_location": "Vị trí cha",
            "is_group": "Là nhóm cha",
            "clinical_area_type": "Khu vực lâm sàng",
            "infection_control_level": "Kiểm soát nhiễm khuẩn",
            "power_backup_available": "Có UPS/máy phát",
            "dept_head": "Người phụ trách",
            "contact_phone": "Số liên hệ",
            "notes": "Ghi chú",
        },
    },
    "IMM Device Model": {
        "name_field": "model_name",
        "export_fields": [
            "name", "model_name", "manufacturer", "asset_category",
            "medical_device_class", "model_version", "country_of_origin",
            "expected_lifespan_years",
            "gmdn_code", "gmdn_term",
            "registration_required", "is_radiation_device",
            "is_pm_required", "pm_interval_days", "pm_alert_days",
            "is_calibration_required", "calibration_interval_days",
            "default_calibration_type",
            "power_supply", "dimensions", "weight_kg",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "model_name": "Tên model",
            "manufacturer": "Nhà sản xuất",
            "asset_category": "Danh mục tài sản",
            "medical_device_class": "Phân loại thiết bị",
            "model_version": "Phiên bản",
            "country_of_origin": "Xuất xứ",
            "expected_lifespan_years": "Tuổi thọ kỳ vọng (năm)",
            "gmdn_code": "Mã GMDN",
            "gmdn_term": "Tên GMDN",
            "registration_required": "Cần số đăng ký BYT",
            "is_radiation_device": "Thiết bị bức xạ",
            "is_pm_required": "Cần bảo trì ĐK",
            "pm_interval_days": "Chu kỳ PM (ngày)",
            "pm_alert_days": "Cảnh báo PM (ngày)",
            "is_calibration_required": "Cần hiệu chuẩn",
            "calibration_interval_days": "Chu kỳ HC (ngày)",
            "default_calibration_type": "Loại hiệu chuẩn",
            "power_supply": "Nguồn điện",
            "dimensions": "Kích thước",
            "weight_kg": "Trọng lượng (kg)",
        },
    },
    "Service Contract": {
        "name_field": "contract_code",
        "export_fields": [
            "name", "contract_code", "contract_title", "supplier",
            "contract_type", "contract_start", "contract_end", "sign_date",
            "contract_value", "auto_renew", "sla_response_hours", "coverage_description",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "contract_code": "Mã hợp đồng",
            "contract_title": "Tên hợp đồng",
            "supplier": "Nhà cung cấp",
            "contract_type": "Loại hợp đồng",
            "contract_start": "Ngày bắt đầu",
            "contract_end": "Ngày kết thúc",
            "sign_date": "Ngày ký",
            "contract_value": "Giá trị (VND)",
            "auto_renew": "Tự động gia hạn",
            "sla_response_hours": "SLA phản hồi (giờ)",
            "coverage_description": "Phạm vi dịch vụ",
        },
    },
    "User": {
        "name_field": "email",
        "export_fields": [
            "email", "full_name", "first_name", "last_name",
            "mobile_no", "ac_department", "imm_approval_status", "roles",
        ],
        "export_labels": {
            "email": "Email đăng nhập",
            "full_name": "Họ và tên",
            "first_name": "Tên",
            "last_name": "Họ",
            "mobile_no": "Điện thoại",
            "ac_department": "Khoa/Phòng",
            "imm_approval_status": "Trạng thái duyệt",
            "roles": "Vai trò (phân cách bằng dấu phẩy)",
        },
    },
    "AC Asset": {
        "name_field": "asset_name",
        "export_fields": [
            "name", "asset_name", "asset_code", "asset_category", "device_model",
            "manufacturer_sn", "udi_code", "byt_reg_no", "byt_reg_expiry",
            "location", "department", "custodian", "responsible_technician", "supplier",
            "purchase_date", "gross_purchase_amount", "warranty_expiry_date",
            "commissioning_date", "in_service_date",
            "depreciation_method", "useful_life_years", "residual_value",
            "total_depreciation_months", "depreciation_frequency", "depreciation_start_date",
            "insurance_policy_no", "insurer_name", "insured_value",
            "insurance_start_date", "insurance_end_date",
            "lifecycle_status", "status", "notes",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "asset_name": "Tên tài sản",
            # ADR-IMM00-ASSETCODE D4: nhãn chuẩn = "Mã tài sản" (PK định danh,
            # để trống = tự sinh). KHÔNG đổi key field — chỉ nhãn VI hiển thị.
            "asset_code": "Mã tài sản",
            "asset_category": "Danh mục tài sản",
            "device_model": "Model thiết bị",
            # ADR-IMM00-ASSETCODE D4: nhãn chuẩn = "Số serial NSX" (field nghiệp
            # vụ riêng, KHÁC Mã tài sản). Khớp AssetCreateView/AssetEditView.
            "manufacturer_sn": "Số serial NSX",
            "udi_code": "Mã UDI",
            "byt_reg_no": "Số ĐKLH Bộ Y tế",
            "byt_reg_expiry": "Hết hạn ĐKLH",
            "location": "Vị trí đặt máy",
            "department": "Khoa phòng quản lý",
            "custodian": "Người phụ trách (email)",
            "responsible_technician": "KTV phụ trách (email)",
            "supplier": "Nhà cung cấp",
            "purchase_date": "Ngày mua",
            "gross_purchase_amount": "Giá mua (VNĐ)",
            "warranty_expiry_date": "Hết hạn bảo hành",
            "commissioning_date": "Ngày nghiệm thu",
            "in_service_date": "Ngày đưa vào sử dụng",
            "depreciation_method": "Phương pháp khấu hao",
            "useful_life_years": "Tuổi thọ hữu ích (năm)",
            "residual_value": "Giá trị thu hồi (VNĐ)",
            "total_depreciation_months": "Tổng tháng KH",
            "depreciation_frequency": "Tần suất KH",
            "depreciation_start_date": "Ngày bắt đầu KH",
            "insurance_policy_no": "Số HĐ bảo hiểm",
            "insurer_name": "Công ty bảo hiểm",
            "insured_value": "Giá trị BH (VNĐ)",
            "insurance_start_date": "Ngày bắt đầu BH",
            "insurance_end_date": "Ngày hết hạn BH",
            "lifecycle_status": "Trạng thái vòng đời",
            "status": "Trạng thái (Asset)",
            "notes": "Ghi chú",
        },
    },
    "AC Supplier": {
        "name_field": "supplier_name",
        "export_fields": [
            "name", "supplier_name", "supplier_code", "vendor_type",
            "country", "tax_id", "email_id", "phone", "mobile_no",
            "support_hotline", "technical_email", "local_representative",
            "website", "address",
            "contract_start", "contract_end", "contract_value",
            "iso_17025_cert", "iso_17025_expiry",
            "iso_13485_cert", "iso_13485_expiry",
            "is_active",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "supplier_name": "Tên nhà cung cấp",
            "supplier_code": "Mã nhà cung cấp",
            "vendor_type": "Loại nhà cung cấp",
            "country": "Quốc gia",
            "tax_id": "Mã số thuế",
            "email_id": "Email liên hệ",
            "phone": "Điện thoại",
            "mobile_no": "Di động",
            "support_hotline": "Hotline hỗ trợ",
            "technical_email": "Email kỹ thuật",
            "local_representative": "Đại diện tại VN",
            "website": "Website",
            "address": "Địa chỉ",
            "contract_start": "Ngày bắt đầu HĐ",
            "contract_end": "Ngày kết thúc HĐ",
            "contract_value": "Giá trị HĐ (VND)",
            "iso_17025_cert": "Chứng chỉ ISO 17025",
            "iso_17025_expiry": "Hạn ISO 17025",
            "iso_13485_cert": "Chứng chỉ ISO 13485",
            "iso_13485_expiry": "Hạn ISO 13485",
            "is_active": "Đang hoạt động",
        },
    },
    # Mẫu bảng kiểm bảo trì — dữ liệu CHA + BẢNG CON (hạng mục kiểm tra).
    # File phẳng: MỖI HÀNG = 1 hạng mục; các cột của mẫu (tên/danh mục/loại/phiên
    # bản/hiệu lực) lặp lại ở mọi hàng cùng mẫu. Nhóm theo (danh mục, loại bảo trì)
    # — đúng khoá định danh của DocType (`autoname: PMCT-{asset_category}-{pm_type}`).
    "PM Checklist Template": {
        "name_field": "template_name",
        "child_table": "checklist_items",
        "group_key_fields": ("asset_category", "pm_type"),
        "parent_fields": [
            "template_name", "asset_category", "pm_type", "version", "effective_date",
        ],
        "child_fields": [
            "description", "measurement_type", "unit",
            "expected_min", "expected_max", "is_critical", "reference_section",
        ],
        "export_fields": [
            "name", "template_name", "asset_category", "pm_type",
            "version", "effective_date", "approved_by",
            "description", "measurement_type", "unit",
            "expected_min", "expected_max", "is_critical", "reference_section",
        ],
        "export_labels": {
            "name": _LABEL_SYSTEM_CODE,
            "template_name": "Tên mẫu bảng kiểm",
            "asset_category": "Danh mục tài sản",
            "pm_type": "Loại bảo trì định kỳ",
            "version": "Phiên bản",
            "effective_date": "Ngày hiệu lực",
            "approved_by": "Người phê duyệt",
            "description": "Nội dung kiểm tra",
            "measurement_type": "Cách ghi nhận kết quả",
            "unit": "Đơn vị đo",
            "expected_min": "Ngưỡng dưới",
            "expected_max": "Ngưỡng trên",
            "is_critical": "Hạng mục trọng yếu",
            "reference_section": "Mục tham chiếu tài liệu",
        },
    },
}

SUPPORTED_REF_DOCTYPES = list(_REF_DATA_CONFIG.keys())

# DocType nhập theo NHÓM (cha + bảng con) thay vì 1 hàng = 1 bản ghi.
GROUPED_IMPORT_DOCTYPES: dict[str, dict] = {
    dt: cfg for dt, cfg in _REF_DATA_CONFIG.items() if cfg.get("child_table")
}

# Cột chỉ có trong template import (không nằm trong export_fields) — vẫn cần nhãn
# tiếng Việt để câu báo lỗi gọi đúng tên cột người dùng nhìn thấy.
_IMPORT_ONLY_LABELS: dict[str, dict[str, str]] = {
    "AC Supplier":      {"supplier_group": "Loại nhà cung cấp"},
    "IMM Device Model": {"specifications": "Thông số kỹ thuật"},
    "Service Contract": {"notes": "Ghi chú"},
}


def field_label(doctype: str, fieldname: str) -> str:
    """Nhãn tiếng Việt của một cột import — dùng trong câu báo lỗi.

    Người dùng chỉ thấy nhãn VI ở hàng 3 của template; báo lỗi kèm fieldname
    tiếng Anh (`asset_category`) buộc họ tự dịch ngược. Fallback về fieldname
    khi cột không có trong cấu hình (file lạ / cột thừa).
    """
    if not fieldname:
        return ""
    cfg = _REF_DATA_CONFIG.get(doctype, {})
    labels: dict[str, str] = dict(cfg.get("export_labels", {}))
    labels.update(_IMPORT_ONLY_LABELS.get(doctype, {}))
    return labels.get(fieldname, fieldname)


def source_row_of(rows: list[dict], row_idx: int) -> int:
    """Số hàng THẬT trong file của dòng dữ liệu thứ `row_idx` (1-based).

    Dòng trống ở giữa file bị parser loại bỏ, nên "dòng thứ 3" của validator
    KHÔNG phải hàng 8 của Excel. Trả về số hàng đã ghi lúc parse; fallback theo
    công thức khi dòng không có dấu vết (file CSV cũ / gọi trực tiếp trong test).
    """
    if 1 <= row_idx <= len(rows):
        recorded = rows[row_idx - 1].get(SOURCE_ROW_KEY)
        if isinstance(recorded, int) and recorded > 0:
            return recorded
    return row_idx + FIRST_DATA_ROW - 1


def enrich_issues(doctype: str, rows: list[dict], issues: list[dict]) -> list[dict]:
    """Bồi `label` (nhãn VI) + `source_row` (hàng thật trong file) vào mỗi lỗi.

    Validator chỉ biết index dòng dữ liệu và fieldname kỹ thuật. FE cần chỉ đúng
    "hàng 12, cột Danh mục tài sản" để người dùng mở file sửa được ngay.
    """
    for issue in issues:
        issue["label"] = field_label(doctype, str(issue.get("field") or ""))
        issue["source_row"] = source_row_of(rows, int(issue.get("row") or 0))
    return issues

# For multi-sheet templates, map each DocType to its sheet name.
# When a user uploads the combined template file, this ensures the correct
# sheet is parsed regardless of which sheet was active when saved.
_SHEET_NAME_MAP: dict[str, str] = {
    "AC Supplier":      "Nhà cung cấp (Supplier)",
    "IMM Device Model": "Mô hình thiết bị (Device Model)",
    "Service Contract": "Hợp đồng (Service Contract)",
    "IMM SLA Policy":   "Chính sách SLA",
    # File mẫu bảng kiểm có thêm sheet "Ví dụ minh hoạ" — khoá tên sheet dữ liệu
    # để người dùng lưu file lúc đang đứng ở sheet ví dụ vẫn nhập đúng sheet.
    "PM Checklist Template": "Bảng kiểm bảo trì",
}


# ─────────────────────────────────────────────────────────────────────────────
# PARSE UPLOADED FILE
# ─────────────────────────────────────────────────────────────────────────────

def parse_upload_file(file_url: str, doctype: str = "") -> tuple[list[str], list[dict]]:
    """
    Parse AssetCore import template (Excel/CSV).

    Template row layout:
        Row 1 — banner (skip)
        Row 2 — fieldnames  ← use as column headers
        Row 3 — VN labels   (skip)
        Row 4 — description (skip)
        Row 5 — example     (skip)
        Row 6+ — data

    For multi-sheet templates the correct sheet is resolved via _SHEET_NAME_MAP.
    Falls back to the active sheet when no mapping is found.

    Returns (fieldnames, rows) where rows is list[dict] keyed by fieldname.
    """
    file_doc = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_doc:
        raise ValueError(f"Không tìm thấy file: {file_url}")
    fdoc = frappe.get_doc("File", file_doc)
    file_path = fdoc.get_full_path()

    if file_url.lower().endswith((".xlsx", ".xls")):
        return _parse_excel(file_path, doctype)
    return _parse_csv(file_path, doctype)


def _parse_excel(file_path: str, doctype: str = "") -> tuple[list[str], list[dict]]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("openpyxl chưa được cài đặt trong bench env") from e

    wb = load_workbook(file_path, data_only=True)

    # Resolve target sheet: prefer mapped name, fall back to active sheet
    target_sheet_name = _SHEET_NAME_MAP.get(doctype, "")
    if target_sheet_name and target_sheet_name in wb.sheetnames:
        ws = wb[target_sheet_name]
    else:
        ws = wb.active

    rows_raw = list(ws.iter_rows(values_only=True))
    if len(rows_raw) < 2:
        raise ValueError("File rỗng hoặc thiếu dòng header (fieldname).")

    # Row index 1 (0-based) = fieldnames — ĐÚNG ở cả hai bố cục
    fieldnames: list[str] = [str(c).strip() if c is not None else "" for c in rows_raw[1]]

    first_data_row = detect_first_data_row(rows_raw, fieldnames, doctype)
    data_rows = rows_raw[first_data_row - 1:]
    return fieldnames, _rows_to_dicts(fieldnames, data_rows, first_data_row)


def _parse_csv(file_path: str, doctype: str = "") -> tuple[list[str], list[dict]]:
    import csv
    with open(file_path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        all_rows = list(reader)

    if len(all_rows) < 2:
        raise ValueError("File CSV rỗng hoặc thiếu dòng header.")

    fieldnames = [c.strip() for c in all_rows[1]]
    rows_raw = [tuple(r) for r in all_rows]
    first_data_row = detect_first_data_row(rows_raw, fieldnames, doctype)
    return fieldnames, _rows_to_dicts(
        fieldnames, rows_raw[first_data_row - 1:], first_data_row,
    )


def _row_matches_label_row(row, fieldnames: list[str], doctype: str) -> bool:
    """Hàng này có phải hàng NHÃN tiếng Việt của template không?

    Tín hiệu dự phòng khi banner bị xoá (người dùng hay xoá hàng 1 cho gọn).
    So sánh bỏ hậu tố " (*)" mà template gắn cho cột bắt buộc.
    """
    if not doctype or not row:
        return False
    hits = comparable = 0
    for value, fieldname in zip(row, fieldnames):
        if not fieldname:
            continue
        label = field_label(doctype, fieldname)
        if not label or label == fieldname:
            continue          # cột không có nhãn riêng ⇒ không phân biệt được
        comparable += 1
        text = str(value or "").strip().removesuffix("(*)").strip()
        if text == label:
            hits += 1
    return comparable > 0 and hits * 2 >= comparable


def detect_first_data_row(rows_raw, fieldnames: list[str], doctype: str = "") -> int:
    """Hàng đầu tiên chứa dữ liệu THẬT của file đang đọc (1-based).

    Hai bố cục cùng tồn tại và phải phân biệt được, nếu không là mất dữ liệu câm:
      - TEMPLATE (5 hàng khung: banner · fieldname · nhãn · mô tả · ví dụ) → hàng 6
      - XUẤT CŨ  (2 hàng khung: nhãn · fieldname)                          → hàng 3

    Hàng fieldname nằm ở hàng 2 trong CẢ HAI, nên nó không phân biệt được; dấu
    hiệu là banner ở A1, dự phòng là hàng nhãn ở hàng 3.
    """
    banner = rows_raw[0][0] if rows_raw and rows_raw[0] else None
    if isinstance(banner, str) and banner.strip().startswith(TEMPLATE_BANNER_PREFIX):
        return FIRST_DATA_ROW
    if len(rows_raw) >= 3 and _row_matches_label_row(rows_raw[2], fieldnames, doctype):
        return FIRST_DATA_ROW
    return LEGACY_EXPORT_FIRST_DATA_ROW


def _normalise_cell(val: Any) -> Any:
    if val is None or (isinstance(val, str) and not val.strip()):
        return ""
    return val.strip() if isinstance(val, str) else val


def _rows_to_dicts(
    fieldnames: list[str], raw_rows, first_data_row: int = FIRST_DATA_ROW,
) -> list[dict]:
    """Chuyển hàng thô → dict theo fieldname, GIỮ số hàng gốc trong file.

    Dòng trống bị loại khỏi kết quả, nên index trong list KHÔNG còn suy ra được
    số hàng Excel — ghi lại ở `SOURCE_ROW_KEY` để báo lỗi chỉ đúng chỗ.
    `first_data_row` khác nhau giữa file mẫu (6) và file xuất bố cục cũ (3), nên
    phải truyền vào chứ không đọc hằng số — sai là báo lỗi chỉ nhầm hàng.
    """
    result = []
    for offset, raw in enumerate(raw_rows):
        row: dict[str, Any] = {
            fn: _normalise_cell(raw[i] if i < len(raw) else None)
            for i, fn in enumerate(fieldnames)
            if fn
        }
        if any(v not in ("", None) for v in row.values()):
            row[SOURCE_ROW_KEY] = first_data_row + offset
            result.append(row)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# BUILD ERROR REPORT
# ─────────────────────────────────────────────────────────────────────────────

def build_error_report(
    fieldnames: list[str], rows: list[dict], errors: list[dict],
    doctype: str = "",
) -> bytes:
    """Return xlsx bytes with error rows highlighted.

    Cột đầu = số hàng THẬT trong file gốc để người dùng mở file lên sửa đúng chỗ;
    ghi chú lỗi gọi cột bằng nhãn tiếng Việt (không phải fieldname tiếng Anh).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
    except ImportError as e:
        raise RuntimeError("openpyxl chưa được cài đặt") from e

    error_map: dict[int, list[str]] = {}
    for e in errors:
        label = field_label(doctype, str(e.get("field") or "")) or "Toàn dòng"
        error_map.setdefault(e["row"], []).append(f"[{label}] {e['message']}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Lỗi import"

    header = ["Hàng trong file"] + fieldnames + ["Trạng thái", "Ghi chú lỗi"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    red_fill = PatternFill("solid", fgColor="FFCCCC")
    ok_fill = PatternFill("solid", fgColor="CCFFCC")

    for i, row in enumerate(rows, start=1):
        data = [row.get(fn, "") for fn in fieldnames]
        notes = error_map.get(i, [])
        status = "Lỗi" if notes else "OK"
        ws.append([source_row_of(rows, i)] + data + [status, "; ".join(notes)])
        xl_row = ws[ws.max_row]
        fill = red_fill if notes else ok_fill
        for cell in xl_row:
            cell.fill = fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT CURRENT DATA
# ─────────────────────────────────────────────────────────────────────────────

_SKIP_USER_ROLES = frozenset({"All", "Guest", "System Manager", "Administrator"})


def _export_users(cfg: dict) -> list[dict]:
    """Fetch User rows including IMM roles from Has Role child table."""
    fields = cfg["export_fields"]
    std_fields = [f for f in fields if f not in ("email", "roles")]
    # Frappe User: name = email; fetch name + standard fields
    frappe_fields = ["name"] + std_fields

    users = frappe.get_all(
        "User",
        filters=[["User", "name", "not in", ["Administrator", "Guest"]]],
        fields=frappe_fields,
        order_by="creation asc",
    )
    if not users:
        return []

    user_names = [u.name for u in users]
    has_roles = frappe.get_all(
        "Has Role",
        filters={"parent": ["in", user_names], "parenttype": "User"},
        fields=["parent", "role"],
    )
    roles_map: dict[str, list[str]] = {}
    for hr in has_roles:
        if hr["role"] not in _SKIP_USER_ROLES:
            roles_map.setdefault(hr["parent"], []).append(hr["role"])

    result: list[dict] = []
    for u in users:
        row: dict = {"email": u.name}
        for f in std_fields:
            row[f] = u.get(f)
        row["roles"] = ", ".join(sorted(roles_map.get(u.name, [])))
        result.append(row)
    return result


def _export_pm_checklist_templates(cfg: dict) -> list[dict]:
    """Trải phẳng mẫu bảng kiểm: mỗi hạng mục con = 1 hàng, cột cha lặp lại.

    Cùng bố cục với file nhập ⇒ xuất ra sửa rồi nhập lại được (mẫu chưa có hạng
    mục nào vẫn ra 1 hàng để không biến mất khỏi file).
    """
    parents = frappe.get_all(
        "PM Checklist Template",
        fields=["name", "template_name", "asset_category", "pm_type",
                "version", "effective_date", "approved_by"],
        order_by="template_name asc",
    )
    if not parents:
        return []

    items = frappe.get_all(
        "PM Checklist Item",
        filters={"parent": ["in", [p["name"] for p in parents]], "parenttype": "PM Checklist Template"},
        fields=["parent", "idx", "description", "measurement_type", "unit",
                "expected_min", "expected_max", "is_critical", "reference_section"],
        order_by="parent asc, idx asc",
    )
    by_parent: dict[str, list[dict]] = {}
    for it in items:
        by_parent.setdefault(it["parent"], []).append(it)

    child_fields = cfg["child_fields"]
    rows: list[dict] = []
    for p in parents:
        base = {k: p.get(k) for k in
                ("name", "template_name", "asset_category", "pm_type",
                 "version", "effective_date", "approved_by")}
        children = by_parent.get(p["name"], [])
        if not children:
            rows.append({**base, **{f: None for f in child_fields}})
            continue
        for it in children:
            rows.append({**base, **{f: it.get(f) for f in child_fields}})
    return rows


def _display_names_for(link_doctype: str, display_field: str, codes: set[str]) -> dict[str, str]:
    """Map mã hệ thống → tên hiển thị cho một tập mã (1 query, không N+1)."""
    if not codes:
        return {}
    found = frappe.get_all(
        link_doctype,
        filters={"name": ["in", sorted(codes)]},
        fields=["name", display_field],
    )
    return {r["name"]: str(r.get(display_field) or "") for r in found if r.get(display_field)}


def resolve_links_to_display(doctype: str, rows: list[dict]) -> list[dict]:
    """Đổi giá trị Link từ mã hệ thống → TÊN hiển thị, tại chỗ.

    File export phải đọc được và import lại được mà không bắt người dùng tra mã:
    cột "Khoa phòng quản lý" phải in "Khoa Hồi sức tích cực", KHÔNG phải
    "AC-DEPT-0007". Mã không tra được (bản ghi đã xoá) giữ nguyên để không mất dấu.
    """
    link_map = LINK_DISPLAY_BY_DOCTYPE.get(doctype, {})
    if not link_map or not rows:
        return rows

    for field, (link_dt, display_field) in link_map.items():
        codes = {str(r.get(field)) for r in rows if r.get(field)}
        lookup = _display_names_for(link_dt, display_field, codes)
        if not lookup:
            continue
        for r in rows:
            val = r.get(field)
            if val and str(val) in lookup:
                r[field] = lookup[str(val)]
    return rows


# Hàng 5 của file xuất — parser BỎ QUA hàng này (vị trí hàng ví dụ của file mẫu).
# Dùng nó để nói thẳng cho người dùng biết dữ liệu bắt đầu từ đâu.
_EXPORT_EXAMPLE_ROW_NOTE = (
    "↓ TỪ HÀNG 6 TRỞ XUỐNG LÀ DỮ LIỆU. Hàng này là hàng ví dụ của khung file — "
    "hệ thống bỏ qua khi nhập lại. Đừng xoá 5 hàng khung ở trên."
)


def _export_banner(doctype: str, title: str) -> str:
    """Banner hàng 1 — vừa hướng dẫn người dùng, vừa là DẤU NHẬN BIẾT bố cục.

    `detect_first_data_row` dựa vào tiền tố này để biết file có 5 hàng khung;
    đổi chuỗi ở đây mà quên hằng `TEMPLATE_BANNER_PREFIX` = nhập lại mất 3 hàng.
    """
    extra = ""
    if doctype in GROUPED_IMPORT_DOCTYPES:
        extra = (
            " MỖI HÀNG = 1 hạng mục; các cột của mẫu lặp lại ở mọi hàng cùng mẫu — "
            "hệ thống gộp theo Danh mục + Loại bảo trì."
        )
    return (
        f"{TEMPLATE_BANNER_PREFIX}: {title} (file xuất từ hệ thống)  |  "
        f"Sửa trực tiếp rồi nhập lại file này. Dữ liệu bắt đầu từ HÀNG {FIRST_DATA_ROW}; "
        f"không xoá/không sửa 5 hàng khung ở trên. "
        f"Cột tham chiếu đã in TÊN — giữ nguyên dạng TÊN, đừng đổi về mã.{extra}"
    )


def _export_column_hint(doctype: str, fieldname: str) -> str:
    """Mô tả cột ở hàng 4 — nói rõ cột nào nhập lại được, cột nào chỉ để đọc."""
    if fieldname == "name":
        return "Mã hệ thống — chỉ để đối chiếu; khi nhập lại hệ thống bỏ qua cột này."
    if fieldname in LINK_DISPLAY_BY_DOCTYPE.get(doctype, {}):
        return "Điền TÊN như trong hệ thống, không điền mã."
    choices = ENUM_DISPLAY_BY_DOCTYPE.get(doctype, {}).get(fieldname, {})
    if choices:
        return "Chọn một trong: " + " / ".join(choices.values())
    return ""


def export_ref_data(doctype: str) -> bytes:
    """Export all records of a ref-data DocType to xlsx bytes.

    Bố cục ghi ra TRÙNG với file mẫu (5 hàng khung) để vòng
    "Xuất Excel → sửa trong Excel → Nhập lại" không mất hàng nào.
    """
    if doctype not in _REF_DATA_CONFIG:
        raise ValueError(f"DocType '{doctype}' không hỗ trợ export")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
    except ImportError as e:
        raise RuntimeError("openpyxl chưa được cài đặt") from e

    cfg = _REF_DATA_CONFIG[doctype]
    fields = cfg["export_fields"]
    labels = cfg["export_labels"]

    if doctype == "User":
        rows = _export_users(cfg)
    elif doctype in GROUPED_IMPORT_DOCTYPES:
        rows = _export_pm_checklist_templates(cfg)
    else:
        rows = frappe.get_all(doctype, fields=fields, order_by="creation asc")

    # Link column phải in TÊN, không in mã — file export cũng là file import lại.
    rows = resolve_links_to_display(doctype, [dict(r) for r in rows])
    # Cột Select in nhãn VI (đúng thứ người dùng thấy trên màn hình).
    for field_name in ENUM_DISPLAY_BY_DOCTYPE.get(doctype, {}):
        for r in rows:
            if r.get(field_name):
                r[field_name] = enum_display(doctype, field_name, str(r[field_name]))

    wb = Workbook()
    ws = wb.active
    ws.title = _SHEET_NAME_MAP.get(doctype, doctype)

    # ── Khung 5 hàng — GIỐNG HỆT file mẫu ────────────────────────────────────
    # File xuất ra chính là file nhập lại. Trước 2026-08-11 export chỉ ghi 2 hàng
    # khung trong khi parser bỏ 5 ⇒ nhập lại mất im lặng 3 bản ghi đầu.
    ws.append([_export_banner(doctype, ws.title)])
    ws.append(fields)                                   # hàng 2 — hợp đồng parser
    ws.append([labels.get(f, f) for f in fields])
    ws.append([_export_column_hint(doctype, f) for f in fields])
    ws.append([_EXPORT_EXAMPLE_ROW_NOTE])

    banner_fill = PatternFill("solid", fgColor="1A5276")
    ws["A1"].fill = banner_fill
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    fn_fill = PatternFill("solid", fgColor="D6EAF8")
    for cell in ws[2]:
        cell.fill = fn_fill
    header_fill = PatternFill("solid", fgColor="2E4053")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
    desc_fill = PatternFill("solid", fgColor="EAF2F8")
    for cell in ws[4]:
        cell.fill = desc_fill
    for cell in ws[5]:
        cell.fill = PatternFill("solid", fgColor="EAFAF1")

    for i, row in enumerate(rows):
        data = [row.get(f) for f in fields]
        ws.append(data)
        # Stripe rows
        if i % 2 == 1:
            stripe = PatternFill("solid", fgColor="F8F9FA")
            for cell in ws[ws.max_row]:
                cell.fill = stripe

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# TEMPLATE PATH
# ─────────────────────────────────────────────────────────────────────────────

_TMPL_IMM00 = "02_imm00_ncc_model_hopdong_sla.xlsx"

_TEMPLATE_MAP: dict[str, str] = {
    "AC Asset Category": "01a_danh_muc_tai_san.xlsx",
    "AC Department":     "01b_khoa_phong.xlsx",
    "AC Location":       "01c_vi_tri.xlsx",
    "AC Supplier":       _TMPL_IMM00,
    "IMM Device Model":  _TMPL_IMM00,
    "Service Contract":  _TMPL_IMM00,
    "IMM SLA Policy":    _TMPL_IMM00,
    "AC Asset":          "03_danh_sach_tai_san.xlsx",
    "AC Spare Part":     "04_danh_sach_phu_tung.xlsx",
    "AC Warehouse":      "05_kho_hang.xlsx",
    "User":              "06_danh_sach_nguoi_dung.xlsx",
    "PM Checklist Template": "07_bang_kiem_bao_tri.xlsx",
}

# Templates nằm ở assetcore/public/import_templates/
# frappe.get_app_path("assetcore") trả về <bench>/apps/assetcore/assetcore
_TEMPLATE_BASE = None


def _get_template_base() -> str:
    global _TEMPLATE_BASE
    if _TEMPLATE_BASE is None:
        import os
        _TEMPLATE_BASE = os.path.join(frappe.get_app_path("assetcore"), "public", "import_templates")
    return _TEMPLATE_BASE


def get_template_path(doctype: str) -> str:
    import os
    filename = _TEMPLATE_MAP.get(doctype)
    if not filename:
        raise ValueError(f"Không có template cho DocType '{doctype}'")
    path = os.path.join(_get_template_base(), filename)
    if not os.path.exists(path):
        raise FileNotFoundError(
            f"Template file không tồn tại: {path}. "
            "Chạy docs/res/imports/generate_templates.py để sinh lại."
        )
    return path


# ─────────────────────────────────────────────────────────────────────────────
# FRAPPE FOLDER MANAGEMENT
# ─────────────────────────────────────────────────────────────────────────────

_IMPORT_FOLDER_ROOT = "Home/AssetCore Imports"


def ensure_import_folder(sub: str = "") -> str:
    """
    Create Frappe folder hierarchy and commit so subsequent upload_file can find it.
    Returns the folder path (e.g. "Home/AssetCore Imports/AC_Department").
    """
    from frappe.core.api.file import create_new_folder

    _create_if_missing(create_new_folder, _IMPORT_FOLDER_ROOT, "Home")
    if not sub:
        frappe.db.commit()
        return _IMPORT_FOLDER_ROOT

    path = f"{_IMPORT_FOLDER_ROOT}/{sub}"
    _create_if_missing(create_new_folder, path, _IMPORT_FOLDER_ROOT)
    frappe.db.commit()
    return path


def _create_if_missing(create_fn, path: str, parent: str) -> None:
    if frappe.db.exists("File", path):
        return
    try:
        create_fn(path.rsplit("/", 1)[-1], parent)
    except Exception:
        frappe.log_error(frappe.get_traceback(), "Import Folder Creation")


# ─────────────────────────────────────────────────────────────────────────────
# VALIDATION HELPER
# ─────────────────────────────────────────────────────────────────────────────

def is_valid_email(value: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value))


def is_valid_gmdn_code(value: str) -> bool:
    """GMDN code: 5 hoặc 6 chữ số."""
    return bool(re.match(r"^\d{5,6}$", value.strip()))
